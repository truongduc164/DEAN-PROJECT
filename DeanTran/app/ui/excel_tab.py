"""
ExcelTab â€“ MINIMAL log-centric UI for Excel translation.

Layout:
  â”Œâ”€ Top bar: language/output/file controls (compact) â”€â”
  â”‚ Progress bars                                       â”‚
  â”‚ LOG panel (~70% height)                             â”‚
  â”‚ Action buttons: Scan | Translate | Pause/Resume/Cancel â”‚
  â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜
"""
from __future__ import annotations

import glob
import logging
import re
import threading
import time
import traceback
from datetime import datetime
from pathlib import Path

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QProgressBar, QTextEdit, QGroupBox, QComboBox, QCheckBox,
    QFileDialog, QFormLayout, QMessageBox, QListWidget, QSpinBox,
    QListWidgetItem, QAbstractItemView, QSplitter, QFrame, QDialog, QLineEdit, QRadioButton, QColorDialog,
)
from PySide6.QtCore import Qt, QThread, Signal, QObject, QTimer
from PySide6.QtGui import QColor

from app.core.signals import signals
from app.settings.settings_manager import settings
from app.ui.dialogs.select_sheets_dialog import SelectSheetsDialog, collect_all_sheets

logger = logging.getLogger("DeanTran.excel_tab")

# â”€â”€ Language options (3Ã—3) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_LANGUAGES = ["Chinese", "English", "Vietnamese"]

_LANG_DISPLAY = {
    "Chinese": "Chinese (中文)",
    "English": "English",
    "Vietnamese": "Vietnamese (Tiếng Việt)",
}
_LANG_REVERSE = {v: k for k, v in _LANG_DISPLAY.items()}

# â”€â”€ Output modes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_OUTPUT_MODES = {
    "Ghi đè (Overwrite)": "overwrite",
    "Dịch trước nguồn (Prefix)": "prefix",
    "Dịch sau nguồn (Suffix)": "suffix",
}
_OUTPUT_REVERSE = {v: k for k, v in _OUTPUT_MODES.items()}


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Translation Worker (runs in QThread)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

class TranslationWorker(QObject):
    """Runs ExcelProcessor.process() in a background thread."""
    log = Signal(str, str)
    progress = Signal(int, int)
    status = Signal(str)
    finished = Signal(str)
    error = Signal(str)
    file_started = Signal(str)
    result_info = Signal(dict)

    def __init__(
        self,
        file_sheets: dict[str, list[str] | None],
        pause_event: threading.Event,
        cancel_event: threading.Event,
    ) -> None:
        super().__init__()
        self.file_sheets = file_sheets
        self._pause_event = pause_event
        self._cancel_event = cancel_event

    def run(self):
        import time
        t_worker_start = time.time()
        try:
            from app.core.translators.translator_service import (
                create_translator, GeminiTranslator, MockTranslator,
            )
            from app.core.event_manager import EventManager
            from app.core.excel_processor import ExcelProcessor, build_output_path
            from app.storage.prompt_store import PromptStore

            # Read settings
            model = settings.get("selected_models.gemini", "gemini-3.1-flash-lite")
            source_lang = settings.get("language_settings.source_lang", "Chinese")
            target_lang = settings.get("language_settings.target_lang", "Vietnamese")
            doc_type = settings.get("document_settings.document_type", "SOP")
            output_mode = settings.get("document_settings.output_mode", "new_sheet")
            max_rows = settings.get("limits_settings.max_rows", 8000)
            max_cols = settings.get("limits_settings.max_cols", 50)
            min_interval = settings.get("speed_settings.min_interval", 0.2)
            check_glossary = settings.get("processing_options.check_glossary_before_ai", True)

            self.log.emit("INFO", f"[config] model={model} {source_lang}->{target_lang} output_mode={output_mode}")
            self.log.emit("INFO", f"[config] max_rows={max_rows} max_cols={max_cols} interval={min_interval}s")

            # Key status
            from app.core.key_provider import KeyProvider
            kp = KeyProvider()
            kp.log_status()
            key = kp.get_key()
            key_loaded = key is not None
            key_prefix = key[:6] + "****" if key and len(key) >= 6 else "(none)"
            self.log.emit("INFO", f"[key] source={kp.key_source} loaded={key_loaded} prefix={key_prefix}")

            # Prompt
            prompt_store = PromptStore()
            prompt = prompt_store.get(doc_type)
            if prompt:
                self.log.emit("INFO", f"[prompt] ({doc_type}): {prompt[:80]}â€¦")

            # Translator
            translator = create_translator(model_name=model)
            self.log.emit("INFO", f"[translator] {type(translator).__name__}")

            total_files = len(self.file_sheets)
            self.log.emit("INFO", f"[files] count={total_files}")
            last_out = ""
            aggregate = {
                "status": "SUCCESS", "translated": 0, "failed": 0,
                "api_calls": 0, "cache_hits": 0, "elapsed": 0.0,
                "output_paths": [],
            }

            for file_idx, (input_path, sel_sheets) in enumerate(self.file_sheets.items(), 1):
                if self._cancel_event.is_set():
                    self.log.emit("WARN", "Cancel requested â€“ skipping remaining files.")
                    break

                self.file_started.emit(Path(input_path).name)
                self.log.emit("INFO", f"â•â•â• File {file_idx}/{total_files}: {Path(input_path).name} â•â•â•")

                em = EventManager()
                em.subscribe("log", lambda lvl, msg: self.log.emit(lvl, msg))
                em.subscribe("progress", lambda cur, tot: self.progress.emit(cur, tot))
                em.subscribe("status", lambda txt: self.status.emit(txt))

                from app.term_engine.glossary_loader import GlossaryLoader
                glossary = None
                if check_glossary:
                    glossary_path = settings.get("processing_options.glossary_file_path", "")
                    if glossary_path and Path(glossary_path).exists():
                        glossary = GlossaryLoader(glossary_path, source_lang=source_lang, target_lang=target_lang)

                proc = ExcelProcessor(
                    translator=translator,
                    event_manager=em,
                    skip_empty=True,
                    skip_numeric=True,
                    source_lang=source_lang,
                    target_lang=target_lang,
                    prompt=prompt,
                    max_rows=max_rows,
                    max_cols=max_cols,
                    min_interval=min_interval,
                    glossary=glossary,
                    use_glossary=check_glossary,
                    batch_mode=True,
                    sheets_to_translate=sel_sheets,
                    pause_event=self._pause_event,
                    cancel_event=self._cancel_event,
                    output_mode=output_mode,
                )

                custom_dir = None
                if settings.get("excel_settings.use_custom_dir", False):
                    custom_dir_path = settings.get("excel_settings.custom_dir_path", "")
                    if custom_dir_path:
                        custom_dir = Path(custom_dir_path)
                        custom_dir.mkdir(parents=True, exist_ok=True)

                out_path = proc.process(input_path, output_dir=custom_dir)
                last_out = str(out_path)
                self.log.emit("INFO", f"[output] {out_path}")

                r = getattr(proc, "last_result", {})
                aggregate["translated"] += r.get("translated", 0)
                aggregate["failed"] += r.get("failed", 0)
                aggregate["api_calls"] += r.get("api_calls", 0)
                aggregate["cache_hits"] += r.get("cache_hits", 0)
                aggregate["elapsed"] += r.get("elapsed", 0.0)
                aggregate["output_paths"].append(str(out_path))

                s = r.get("status", "SUCCESS")
                if s == "FAILED":
                    aggregate["status"] = "FAILED"
                elif s in ("PARTIAL", "CANCELLED") and aggregate["status"] == "SUCCESS":
                    aggregate["status"] = s

            if self._cancel_event.is_set() and aggregate["status"] == "SUCCESS":
                aggregate["status"] = "CANCELLED"

            aggregate["elapsed"] = time.time() - t_worker_start
            self.result_info.emit(aggregate)
            self.finished.emit(last_out)

        except Exception as exc:
            tb = traceback.format_exc()
            self.log.emit("ERROR", f"Translation failed: {type(exc).__name__}: {exc}")
            logger.error("Translation worker exception:\n%s", tb)
            self.error.emit(str(exc))
            self.result_info.emit({"status": "FAILED", "error": str(exc)})
            self.finished.emit("")


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ExcelTab â€“ Minimal, log-centric UI
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

class ExcelTab(QWidget):
    """Minimal log-centric Excel translation tab."""

    def __init__(self):
        super().__init__()
        self._thread: QThread | None = None
        self._worker: TranslationWorker | None = None
        self._pause_event = threading.Event()
        self._pause_event.set()
        self._cancel_event = threading.Event()
        self._last_result: dict = {}
        self._file_sheets: dict[str, list[str] | None] = {}

        self._build_ui()
        self._load_from_settings()
        signals.log_signal.connect(self.append_log)

        # Clock timer
        self._clock_timer = QTimer(self)
        self._clock_timer.timeout.connect(self._update_clock)
        self._clock_timer.start(1000)
        self._update_clock()

    def _build_ui(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(12)

        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(6)
        left_widget.setMaximumWidth(450)

        left_layout.addWidget(QLabel("<b>1. Nhap File Excel:</b>"))
        row_files_btn = QHBoxLayout()
        self.btn_add = QPushButton("+ Add File")
        self.btn_add.clicked.connect(self._on_add_files)
        self.btn_remove = QPushButton("Remove")
        self.btn_remove.clicked.connect(self._on_remove_file)
        self.btn_clear = QPushButton("Clear")
        self.btn_clear.clicked.connect(self._on_clear_files)
        row_files_btn.addWidget(self.btn_add)
        row_files_btn.addWidget(self.btn_remove)
        row_files_btn.addWidget(self.btn_clear)
        row_files_btn.addStretch()
        left_layout.addLayout(row_files_btn)

        self.lst_files = QListWidget()
        self.lst_files.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.lst_files.setMaximumHeight(110)
        self.lst_files.setStyleSheet("font-size: 12px;")
        left_layout.addWidget(self.lst_files)

        options_group = QGroupBox("2. Chon tinh nang")
        options_form = QFormLayout()
        self.options_form = options_form
        options_form.setContentsMargins(8, 4, 8, 4)

        self.cb_source = QComboBox()
        self.cb_source.addItems([_LANG_DISPLAY[l] for l in _LANGUAGES])
        self.cb_source.currentTextChanged.connect(self._on_source_changed)
        options_form.addRow("Source:", self.cb_source)

        self.cb_target = QComboBox()
        self.cb_target.addItems([_LANG_DISPLAY[l] for l in _LANGUAGES])
        self.cb_target.currentTextChanged.connect(self._on_target_changed)
        options_form.addRow("Target:", self.cb_target)

        self.cb_output_mode = QComboBox()
        self.cb_output_mode.addItems(list(_OUTPUT_MODES.keys()))
        self.cb_output_mode.currentTextChanged.connect(self._on_output_mode_changed)
        options_form.addRow("Output:", self.cb_output_mode)

        # Prompt selector
        prompt_row = QHBoxLayout()
        self.cb_prompt = QComboBox()
        self.cb_prompt.setEditable(True)
        self._refresh_prompt_list()
        self.cb_prompt.currentTextChanged.connect(
            lambda t: settings.set("document_settings.document_type", t)
        )
        self.btn_edit_prompt = QPushButton("✏️")
        self.btn_edit_prompt.setMaximumWidth(32)
        self.btn_edit_prompt.setToolTip("Chỉnh sửa prompt")
        self.btn_edit_prompt.clicked.connect(self._on_edit_prompt)
        prompt_row.addWidget(self.cb_prompt, stretch=1)
        prompt_row.addWidget(self.btn_edit_prompt)
        options_form.addRow("Prompt:", prompt_row)

        self.chk_glossary = QCheckBox("Lọc từ điển trước (Glossary)")
        self.chk_glossary.setChecked(settings.get("processing_options.check_glossary_before_ai", True))
        self.chk_glossary.stateChanged.connect(self._on_glossary_toggled)
        options_form.addRow("", self.chk_glossary)

        glossary_row = QHBoxLayout()
        self.le_glossary_path = QLineEdit(settings.get("processing_options.glossary_file_path", ""))
        self.le_glossary_path.setPlaceholderText("File từ điển (.xlsx, .json)")
        self.le_glossary_path.editingFinished.connect(
            lambda: settings.set("processing_options.glossary_file_path", self.le_glossary_path.text().strip())
        )
        self.btn_browse_glossary = QPushButton("Browse...")
        self.btn_browse_glossary.clicked.connect(self._browse_glossary)
        self.btn_view_glossary = QPushButton("👁")
        self.btn_view_glossary.setMaximumWidth(32)
        self.btn_view_glossary.setToolTip("Xem từ điển")
        self.btn_view_glossary.clicked.connect(self._view_glossary)
        glossary_row.addWidget(self.le_glossary_path)
        glossary_row.addWidget(self.btn_browse_glossary)
        glossary_row.addWidget(self.btn_view_glossary)
        options_form.addRow("Từ điển:", glossary_row)

        self.chk_textbox = QCheckBox("Dich TextBox")
        self.chk_textbox.setChecked(settings.get("processing_options.translate_excel_textboxes", False))
        self.chk_textbox.stateChanged.connect(
            lambda s: settings.set("processing_options.translate_excel_textboxes", s == Qt.Checked.value)
        )
        options_form.addRow("", self.chk_textbox)

        self.chk_ocr = QCheckBox("Dich Anh (OCR)")
        self.chk_ocr.setChecked(settings.get("ocr_settings.image_text_translation_enabled", False))
        self.chk_ocr.stateChanged.connect(self._on_ocr_toggled)
        options_form.addRow("", self.chk_ocr)

        self.cb_ocr_engine = QComboBox()
        self.cb_ocr_engine.addItems(["paddle", "google_vision"])
        self.cb_ocr_engine.setCurrentText(settings.get("ocr_settings.engine", "paddle"))
        self.cb_ocr_engine.currentTextChanged.connect(lambda t: settings.set("ocr_settings.engine", t))
        options_form.addRow("OCR Engine:", self.cb_ocr_engine)

        self.chk_auto_open = QCheckBox("Tu dong mo file")
        self.chk_auto_open.setChecked(settings.get("processing_options.auto_open_file", False))
        self.chk_auto_open.stateChanged.connect(
            lambda s: settings.set("processing_options.auto_open_file", s == Qt.Checked.value)
        )
        options_form.addRow("", self.chk_auto_open)

        # Custom output directory
        self.chk_custom_dir = QCheckBox("Lưu vào thư mục khác (Custom Output)")
        self.chk_custom_dir.stateChanged.connect(self._on_custom_dir_toggled)

        custom_dir_row = QHBoxLayout()
        self.le_custom_dir = QLineEdit()
        self.le_custom_dir.setPlaceholderText("Mặc định: cùng thư mục file gốc")
        self.le_custom_dir.setReadOnly(True)
        self.btn_browse_custom_dir = QPushButton("Browse...")
        self.btn_browse_custom_dir.clicked.connect(self._browse_custom_dir)
        custom_dir_row.addWidget(self.le_custom_dir)
        custom_dir_row.addWidget(self.btn_browse_custom_dir)

        options_form.addRow(self.chk_custom_dir)
        options_form.addRow("Thư mục lưu:", custom_dir_row)

        self.chk_pinyin = QCheckBox("Show Pinyin below Chinese text")
        self.chk_pinyin.setChecked(settings.get("processing_options.add_chinese_pinyin", False))
        self.chk_pinyin.stateChanged.connect(
            lambda s: settings.set("processing_options.add_chinese_pinyin", s == Qt.Checked.value)
        )
        options_form.addRow("", self.chk_pinyin)

        pinyin_style_row = QHBoxLayout()
        pinyin_style_row.addWidget(QLabel("Pinyin:"))
        self.cb_pinyin_font = QComboBox()
        self.cb_pinyin_font.setEditable(True)
        self.cb_pinyin_font.addItems(["Arial", "Times New Roman", "Calibri", "Tahoma"])
        self.cb_pinyin_font.setCurrentText(settings.get("processing_options.pinyin_font_family", "Arial"))
        self.cb_pinyin_font.setMaximumWidth(120)
        self.cb_pinyin_font.setToolTip("Pinyin font family")
        pinyin_style_row.addWidget(self.cb_pinyin_font)
        self.sp_pinyin_size = QSpinBox()
        self.sp_pinyin_size.setRange(6, 48)
        self.sp_pinyin_size.setValue(settings.get("processing_options.pinyin_font_size", 10))
        self.sp_pinyin_size.setToolTip("Pinyin font size (pt)")
        self.sp_pinyin_size.setMaximumWidth(55)
        pinyin_style_row.addWidget(self.sp_pinyin_size)
        self.le_pinyin_color = QLineEdit(settings.get("processing_options.pinyin_font_color", "#888888"))
        self.le_pinyin_color.setMaximumWidth(75)
        self.le_pinyin_color.setToolTip("Pinyin font color (hex)")
        pinyin_style_row.addWidget(self.le_pinyin_color)
        self.btn_pinyin_color = QPushButton("🎨")
        self.btn_pinyin_color.setMaximumWidth(28)
        self.btn_pinyin_color.setToolTip("Pick pinyin color")
        self.btn_pinyin_color.clicked.connect(self._pick_pinyin_color)
        pinyin_style_row.addWidget(self.btn_pinyin_color)
        self.chk_pinyin_keep_format = QCheckBox("Keep original format")
        self.chk_pinyin_keep_format.setChecked(settings.get("processing_options.pinyin_format_mode", "custom") == "keep_original")
        self.chk_pinyin_keep_format.setToolTip("Keep original font/size/color for pinyin")
        self.chk_pinyin_keep_format.stateChanged.connect(
            lambda s: settings.set("processing_options.pinyin_format_mode", "keep_original" if s == Qt.Checked.value else "custom")
        )
        pinyin_style_row.addWidget(self.chk_pinyin_keep_format)
        self.cb_pinyin_font.currentTextChanged.connect(
            lambda t: settings.set("processing_options.pinyin_font_family", t)
        )
        self.sp_pinyin_size.valueChanged.connect(
            lambda v: settings.set("processing_options.pinyin_font_size", v)
        )
        self.le_pinyin_color.editingFinished.connect(
            lambda: settings.set("processing_options.pinyin_font_color", self._normalize_hex_color(self.le_pinyin_color.text()))
        )
        pinyin_style_row.addStretch()
        options_form.addRow("", pinyin_style_row)

        options_group.setLayout(options_form)
        left_layout.addWidget(options_group)

        self.grp_style = QGroupBox("3. Translated Text Format")
        style_layout = QVBoxLayout()
        style_layout.setContentsMargins(8, 4, 8, 4)

        self.rbtn_keep_format = QRadioButton("Keep original format")
        self.rbtn_keep_format.toggled.connect(self._toggle_custom_format)
        style_layout.addWidget(self.rbtn_keep_format)

        self.rbtn_custom_format = QRadioButton("Custom translated format")
        style_layout.addWidget(self.rbtn_custom_format)

        self.custom_format_widget = QWidget()
        custom_layout = QFormLayout(self.custom_format_widget)
        custom_layout.setContentsMargins(16, 0, 0, 0)

        self.cb_font = QComboBox()
        self.cb_font.setEditable(True)
        self.cb_font.addItems(["Mặc định", "Arial", "Times New Roman", "Calibri", "Tahoma"])
        custom_layout.addRow("Font:", self.cb_font)

        size_row = QHBoxLayout()
        self.chk_default_size = QCheckBox("Mặc định")
        self.sp_size = QSpinBox()
        self.sp_size.setRange(8, 72)
        self.sp_size.setValue(14)
        self.chk_default_size.toggled.connect(self._on_default_size_toggled)
        size_row.addWidget(self.chk_default_size)
        size_row.addWidget(self.sp_size)
        size_row.addStretch()
        custom_layout.addRow("Size:", size_row)

        color_row = QHBoxLayout()
        self.chk_default_color = QCheckBox("Mặc định")
        self.le_color = QLineEdit("#000000")
        self.le_color.setMaximumWidth(90)
        self.le_color.editingFinished.connect(self._on_color_text_changed)
        self.btn_color_pick = QPushButton("...")
        self.btn_color_pick.setMaximumWidth(32)
        self.btn_color_pick.clicked.connect(self._pick_color)
        self.btn_color_preview = QPushButton("")
        self.btn_color_preview.setMaximumWidth(32)
        self.btn_color_preview.setEnabled(False)
        
        self.chk_default_color.toggled.connect(self._on_default_color_toggled)
        
        color_row.addWidget(self.chk_default_color)
        color_row.addWidget(self.le_color)
        color_row.addWidget(self.btn_color_pick)
        color_row.addWidget(self.btn_color_preview)
        color_row.addStretch()
        custom_layout.addRow("Color:", color_row)

        self.chk_bold = QCheckBox("Bold")
        self.chk_bold.setTristate(True)
        custom_layout.addRow("", self.chk_bold)
        self.chk_italic = QCheckBox("Italic")
        self.chk_italic.setTristate(True)
        custom_layout.addRow("", self.chk_italic)
        self.chk_underline = QCheckBox("Underline")
        self.chk_underline.setTristate(True)
        custom_layout.addRow("", self.chk_underline)

        style_layout.addWidget(self.custom_format_widget)
        self.grp_style.setLayout(style_layout)
        left_layout.addWidget(self.grp_style)
        left_layout.addStretch()
        root.addWidget(left_widget)

        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(6)

        row_time = QHBoxLayout()
        self.lbl_datetime = QLabel("")
        self.lbl_datetime.setStyleSheet("color: #888; font-size: 12px;")
        self.lbl_warning = QLabel("")
        self.lbl_warning.setStyleSheet("color: red; font-weight: bold;")
        row_time.addWidget(self.lbl_datetime)
        row_time.addStretch()
        row_time.addWidget(self.lbl_warning)
        right_layout.addLayout(row_time)

        self.lbl_current_file = QLabel("")
        self.lbl_current_file.setStyleSheet("font-size: 12px; color: #555;")
        right_layout.addWidget(self.lbl_current_file)

        self.bar_progress = QProgressBar()
        self.bar_progress.setFormat("%p% (%v / %m cells)")
        self.bar_progress.setMinimumHeight(22)
        right_layout.addWidget(self.bar_progress)

        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setStyleSheet(
            "background-color: #1e1e2e; color: #cdd6f4; "
            "font-family: 'Cascadia Code', 'Consolas', monospace; "
            "font-size: 12px; padding: 6px; border-radius: 4px;"
        )
        right_layout.addWidget(self.txt_log, stretch=7)

        self.lbl_status = QLabel("")
        self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px;")
        right_layout.addWidget(self.lbl_status)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        self.btn_translate = QPushButton("Translate Excel")
        self.btn_translate.setStyleSheet(
            "padding: 10px 20px; font-weight: bold; font-size: 14px; "
            "background-color: #89b4fa; color: #1e1e2e; border-radius: 6px;"
        )
        self.btn_translate.clicked.connect(self._on_translate_clicked)

        self.btn_pause = QPushButton("Pause")
        self.btn_resume = QPushButton("Resume")
        self.btn_cancel = QPushButton("Cancel")
        self.btn_retry = QPushButton("Retry Failed")
        for btn in (self.btn_pause, self.btn_resume, self.btn_cancel, self.btn_retry):
            btn.setStyleSheet("padding: 8px 14px; font-weight: bold;")
            btn.setEnabled(False)

        self.btn_pause.clicked.connect(self._on_pause)
        self.btn_resume.clicked.connect(self._on_resume)
        self.btn_cancel.clicked.connect(self._on_cancel)
        self.btn_retry.clicked.connect(self._on_retry_failed)

        btn_row.addWidget(self.btn_translate)
        btn_row.addWidget(self.btn_pause)
        btn_row.addWidget(self.btn_resume)
        btn_row.addWidget(self.btn_cancel)
        btn_row.addWidget(self.btn_retry)
        btn_row.addStretch()
        right_layout.addLayout(btn_row)

        root.addWidget(right_widget, stretch=1)

    # Settings sync
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    @staticmethod
    def _set_form_row_visible(form_layout: QFormLayout, field_widget: QWidget, visible: bool):
        label = form_layout.labelForField(field_widget)
        if label is not None:
            label.setVisible(visible)
        field_widget.setVisible(visible)

    def _update_ocr_ui_visibility(self, enabled: bool | None = None):
        if enabled is None:
            enabled = self.chk_ocr.isChecked()
        self._set_form_row_visible(self.options_form, self.cb_ocr_engine, enabled)

    def _update_glossary_ui_visibility(self, enabled: bool | None = None):
        if enabled is None:
            enabled = self.chk_glossary.isChecked()
        self.le_glossary_path.setEnabled(enabled)
        self.btn_browse_glossary.setEnabled(enabled)
        self.btn_view_glossary.setEnabled(enabled)

    def _on_glossary_toggled(self, state):
        enabled = state == Qt.Checked.value
        settings.set("processing_options.check_glossary_before_ai", enabled)
        self._update_glossary_ui_visibility(enabled)

    def _browse_glossary(self):
        fname, _ = QFileDialog.getOpenFileName(
            self, "Select Glossary File", "", "Excel/JSON (*.xlsx *.xls *.json)"
        )
        if fname:
            self.le_glossary_path.setText(fname)
            settings.set("processing_options.glossary_file_path", fname)

    def _view_glossary(self):
        path = self.le_glossary_path.text().strip()
        if not path:
            QMessageBox.information(self, "Thông báo", "Chưa chọn file từ điển.")
            return
        from app.ui.dialogs.glossary_viewer_dialog import GlossaryViewerDialog
        dlg = GlossaryViewerDialog(path, parent=self)
        dlg.exec()

    def _on_custom_dir_toggled(self, state=None):
        enabled = self.chk_custom_dir.isChecked()
        settings.set("excel_settings.use_custom_dir", enabled)
        self.le_custom_dir.setEnabled(enabled)
        self.btn_browse_custom_dir.setEnabled(enabled)

    def _browse_custom_dir(self):
        curr_dir = self.le_custom_dir.text().strip() or settings.get("file_picker.last_export_dir", "")
        dir_path = QFileDialog.getExistingDirectory(self, "Chọn thư mục lưu kết quả", curr_dir)
        if dir_path:
            self.le_custom_dir.setText(dir_path)
            settings.set("excel_settings.custom_dir_path", dir_path)

    def _refresh_prompt_list(self):
        from app.storage.prompt_store import PromptStore
        store = PromptStore()
        modes = store.modes()
        current = settings.get("document_settings.document_type", "General")
        self.cb_prompt.blockSignals(True)
        self.cb_prompt.clear()
        self.cb_prompt.addItems(modes)
        idx = self.cb_prompt.findText(current)
        if idx >= 0:
            self.cb_prompt.setCurrentIndex(idx)
        else:
            self.cb_prompt.setCurrentText(current)
        self.cb_prompt.blockSignals(False)

    def _on_edit_prompt(self):
        mode = self.cb_prompt.currentText().strip()
        if not mode:
            return
        from app.ui.dialogs.prompt_edit_dialog import PromptEditDialog
        dlg = PromptEditDialog(mode, parent=self)
        if dlg.exec():
            self._refresh_prompt_list()

    def _on_ocr_toggled(self, state):
        enabled = state == Qt.Checked.value
        settings.set("ocr_settings.image_text_translation_enabled", enabled)
        self._update_ocr_ui_visibility(enabled)

    def _toggle_custom_format(self):
        self.custom_format_widget.setEnabled(self.rbtn_custom_format.isChecked())

    def _normalize_hex_color(self, value: str) -> str:
        text = (value or "").strip().upper()
        if not text:
            return "#000000"
        if not text.startswith("#"):
            text = f"#{text}"
        if re.fullmatch(r"#[0-9A-F]{6}", text):
            return text
        return "#000000"

    def _refresh_color_preview(self):
        color = self._normalize_hex_color(self.le_color.text())
        self.btn_color_preview.setStyleSheet(
            f"background-color: {color}; border: 1px solid #888; border-radius: 3px;"
        )

    def _on_color_text_changed(self):
        normalized = self._normalize_hex_color(self.le_color.text())
        self.le_color.setText(normalized)
        self._refresh_color_preview()

    def _on_default_color_toggled(self, checked):
        self.le_color.setEnabled(not checked)
        self.btn_color_pick.setEnabled(not checked)
        self.btn_color_preview.setVisible(not checked)

    def _on_default_size_toggled(self, checked):
        self.sp_size.setEnabled(not checked)

    def _pick_color(self):
        current = QColor(self._normalize_hex_color(self.le_color.text()))
        picked = QColorDialog.getColor(current, self, "Chon mau chu")
        if not picked.isValid():
            return
        self.le_color.setText(picked.name().upper())
        self._refresh_color_preview()

    def _load_from_settings(self):
        for w in (self.cb_source, self.cb_target, self.cb_output_mode, self.cb_ocr_engine):
            w.blockSignals(True)

        src = settings.get("language_settings.source_lang", "Chinese")
        self.cb_source.setCurrentText(_LANG_DISPLAY.get(src, src))

        tgt = settings.get("language_settings.target_lang", "Vietnamese")
        self.cb_target.setCurrentText(_LANG_DISPLAY.get(tgt, tgt))

        om = settings.get("document_settings.output_mode", "overwrite")
        self.cb_output_mode.setCurrentText(_OUTPUT_REVERSE.get(om, "Ghi đè (Overwrite)"))

        ocr = settings.get("ocr_settings.image_text_translation_enabled", False)
        self.chk_ocr.setChecked(ocr)
        self.cb_ocr_engine.setCurrentText(settings.get("ocr_settings.engine", "paddle"))

        glossary = settings.get("processing_options.check_glossary_before_ai", True)
        self.chk_glossary.setChecked(glossary)
        self.le_glossary_path.setText(settings.get("processing_options.glossary_file_path", ""))

        auto_open = settings.get("processing_options.auto_open_file", False)
        self.chk_auto_open.setChecked(auto_open)
        self.chk_pinyin.setChecked(settings.get("processing_options.add_chinese_pinyin", False))

        fmt_mode = settings.get(
            "text_style_settings.translated_text_format_mode",
            "keep_original_format" if settings.get("text_style_settings.keep_format", True) else "custom_format",
        )
        is_keep = fmt_mode == "keep_original_format"
        self.rbtn_keep_format.setChecked(is_keep)
        self.rbtn_custom_format.setChecked(not is_keep)
        self.cb_font.setCurrentText(settings.get("text_style_settings.font_family", "Arial"))
        font_size = settings.get("text_style_settings.font_size", 14)
        if font_size == 0:
            self.chk_default_size.setChecked(True)
            self.sp_size.setValue(14)
        else:
            self.chk_default_size.setChecked(False)
            self.sp_size.setValue(font_size)
        
        font_color = settings.get("text_style_settings.font_color", "#000000")
        if font_color == "default":
            self.chk_default_color.setChecked(True)
            self.le_color.setText("#000000")
        else:
            self.chk_default_color.setChecked(False)
            self.le_color.setText(self._normalize_hex_color(font_color))
        def _set_tri_state(chk, val):
            if val == "default":
                chk.setCheckState(Qt.PartiallyChecked)
            else:
                chk.setCheckState(Qt.Checked if val else Qt.Unchecked)

        _set_tri_state(self.chk_bold, settings.get("text_style_settings.bold", "default"))
        _set_tri_state(self.chk_italic, settings.get("text_style_settings.italic", "default"))
        _set_tri_state(self.chk_underline, settings.get("text_style_settings.underline", "default"))
        use_custom_dir = settings.get("excel_settings.use_custom_dir", False)
        self.chk_custom_dir.setChecked(use_custom_dir)
        self.le_custom_dir.setText(settings.get("excel_settings.custom_dir_path", ""))
        self._on_custom_dir_toggled(Qt.Checked if use_custom_dir else Qt.Unchecked)

        self._refresh_color_preview()
        self._toggle_custom_format()

        for w in (
            self.cb_source,
            self.cb_target,
            self.cb_output_mode,
            self.cb_ocr_engine,
            self.chk_ocr,
            self.chk_auto_open,
            self.chk_pinyin,
        ):
            w.blockSignals(False)

        self._update_ocr_ui_visibility(ocr)
        self._update_glossary_ui_visibility(glossary)
        self._check_same_language()

    @staticmethod
    def _normalize_hex_color(text: str) -> str:
        """Normalize a hex color string."""
        text = text.strip().upper()
        if not text.startswith("#"):
            text = f"#{text}"
        if re.fullmatch(r"#[0-9A-F]{6}", text):
            return text
        return "#000000"

    def _pick_pinyin_color(self):
        """Open color dialog for pinyin font color."""
        current = QColor(self._normalize_hex_color(self.le_pinyin_color.text()))
        picked = QColorDialog.getColor(current, self, "Chọn màu pinyin")
        if not picked.isValid():
            return
        self.le_pinyin_color.setText(picked.name().upper())

    def _update_clock(self):
        self.lbl_datetime.setText(datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))

    def _on_source_changed(self, text: str):
        lang = _LANG_REVERSE.get(text, text)
        settings.set("language_settings.source_lang", lang)
        self._check_same_language()

    def _on_target_changed(self, text: str):
        lang = _LANG_REVERSE.get(text, text)
        settings.set("language_settings.target_lang", lang)
        self._check_same_language()

    def _on_output_mode_changed(self, text: str):
        mode = _OUTPUT_MODES.get(text, "new_sheet")
        settings.set("document_settings.output_mode", mode)

    def _check_same_language(self):
        src = settings.get("language_settings.source_lang", "Chinese")
        tgt = settings.get("language_settings.target_lang", "Vietnamese")
        if src == tgt:
            self.lbl_warning.setText("Source = Target!")
            self.btn_translate.setEnabled(False)
        else:
            self.lbl_warning.setText("")
            self._update_translate_button()

    def _update_translate_button(self):
        """Enable Translate only if files selected and languages differ."""
        src = settings.get("language_settings.source_lang", "Chinese")
        tgt = settings.get("language_settings.target_lang", "Vietnamese")
        has_files = len(self._file_sheets) > 0
        lang_ok = src != tgt
        self.btn_translate.setEnabled(has_files and lang_ok)

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # File management
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    def _on_add_files(self):
        fnames, _ = QFileDialog.getOpenFileNames(
            self, "Select Excel files", "", "Excel Files (*.xlsx *.xls)"
        )
        if not fnames:
            return
            
        converted_fnames = []
        for fname in fnames:
            if fname.lower().endswith(".xls"):
                self.append_log("INFO", f"Converting {Path(fname).name} to .xlsx...")
                try:
                    from PySide6.QtWidgets import QApplication
                    QApplication.processEvents()
                    import win32com.client
                    import pythoncom
                    pythoncom.CoInitialize()
                    excel = win32com.client.DispatchEx("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    wb = excel.Workbooks.Open(str(Path(fname).resolve()))
                    new_fname = str(Path(fname).with_suffix(".xlsx").resolve())
                    wb.SaveAs(new_fname, FileFormat=51)
                    wb.Close()
                    excel.Quit()
                    pythoncom.CoUninitialize()
                    converted_fnames.append(new_fname)
                    self.append_log("INFO", f"Converted successfully: {Path(new_fname).name}")
                except Exception as e:
                    self.append_log("ERROR", f"Failed to convert {Path(fname).name}: {e}")
            else:
                converted_fnames.append(fname)
                
        for fname in converted_fnames:
            if fname in self._file_sheets:
                continue
            self._file_sheets[fname] = None  # None = all sheets
            item = QListWidgetItem(Path(fname).name)
            item.setData(Qt.UserRole, fname)
            self.lst_files.addItem(item)
            self.append_log("INFO", f"[file_loaded] {Path(fname).name}")

        self.append_log("INFO", f"[files] count={len(self._file_sheets)}")
        self._update_translate_button()

    def _on_remove_file(self):
        for item in self.lst_files.selectedItems():
            fname = item.data(Qt.UserRole)
            if fname in self._file_sheets:
                del self._file_sheets[fname]
            self.lst_files.takeItem(self.lst_files.row(item))
        self._update_translate_button()

    def _on_clear_files(self):
        self.lst_files.clear()
        self._file_sheets = {}
        self.append_log("INFO", "Cleared all files.")
        self._update_translate_button()

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Pause / Resume / Cancel
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    def _on_pause(self):
        self._pause_event.clear()
        self.btn_pause.setEnabled(False)
        self.btn_resume.setEnabled(True)
        self.append_log("INFO", "Paused - no new API calls.")
        self.lbl_status.setText("Paused")
        self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px; color: orange; font-weight: bold;")

    def _on_resume(self):
        self._pause_event.set()
        self.btn_pause.setEnabled(True)
        self.btn_resume.setEnabled(False)
        self.append_log("INFO", "Resumed.")
        self.lbl_status.setText("Running...")
        self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px; color: #89b4fa; font-weight: bold;")

    def _on_cancel(self):
        self._cancel_event.set()
        self.btn_pause.setEnabled(False)
        self.btn_resume.setEnabled(False)
        self.btn_cancel.setEnabled(False)
        self.append_log("WARN", "Cancel requested...")

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Translate
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    def _on_translate_clicked(self):
        if not self._file_sheets:
            QMessageBox.warning(self, "Error", "Add at least one file.")
            return

        fmt_mode = "keep_original_format" if self.rbtn_keep_format.isChecked() else "custom_format"
        settings.set("text_style_settings.translated_text_format_mode", fmt_mode)
        settings.set("text_style_settings.keep_format", self.rbtn_keep_format.isChecked())
        settings.set("text_style_settings.font_family", self.cb_font.currentText())
        settings.set("text_style_settings.font_size", 0 if self.chk_default_size.isChecked() else self.sp_size.value())
        if self.chk_default_color.isChecked():
            settings.set("text_style_settings.font_color", "default")
        else:
            settings.set("text_style_settings.font_color", self._normalize_hex_color(self.le_color.text()))
        def _get_tri_state(chk):
            state = chk.checkState()
            if state == Qt.PartiallyChecked:
                return "default"
            return state == Qt.Checked

        settings.set("text_style_settings.bold", _get_tri_state(self.chk_bold))
        settings.set("text_style_settings.italic", _get_tri_state(self.chk_italic))
        settings.set("text_style_settings.underline", _get_tri_state(self.chk_underline))
        settings.set("processing_options.add_chinese_pinyin", self.chk_pinyin.isChecked())

        src = settings.get("language_settings.source_lang", "Chinese")
        tgt = settings.get("language_settings.target_lang", "Vietnamese")
        if src == tgt:
            QMessageBox.warning(self, "Error", "Source and Target language must be different.")
            return

        for fp in self._file_sheets:
            if not Path(fp).exists():
                QMessageBox.warning(self, "Error", f"File not found: {fp}")
                return

        if self._thread and self._thread.isRunning():
            QMessageBox.warning(self, "Busy", "Translation already running.")
            return

        # â”€â”€ Open sheet selection dialog â”€â”€
        all_sheets = collect_all_sheets(list(self._file_sheets.keys()))
        dlg = SelectSheetsDialog(all_sheets, parent=self)
        if dlg.exec() != QDialog.Accepted:
            return  # User cancelled

        selection = dlg.get_selection()
        if selection is None:
            return

        # Merge selection into _file_sheets
        for fp in self._file_sheets:
            if fp in selection:
                self._file_sheets[fp] = selection[fp]

        # Count total selected
        total_sheets = sum(
            len(s) if s else 0 for s in self._file_sheets.values()
        )
        self.append_log("INFO", f"[sheets_selected] count={total_sheets}")
        for fp, sheets in self._file_sheets.items():
            self.append_log("INFO", f"  {Path(fp).name}: {sheets}")

        # Reset
        self._pause_event.set()
        self._cancel_event.clear()
        self.txt_log.clear()
        self.bar_progress.setValue(0)

        output_mode = settings.get("document_settings.output_mode", "new_sheet")
        self.append_log("INFO", f"[start] files={len(self._file_sheets)} {src}->{tgt} mode={output_mode}")

        # UI state
        self.btn_translate.setEnabled(False)
        self.btn_pause.setEnabled(True)
        self.btn_resume.setEnabled(False)
        self.btn_cancel.setEnabled(True)
        self.btn_retry.setEnabled(False)
        self.lbl_status.setText("Running...")
        self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px; color: #89b4fa; font-weight: bold;")

        # Worker
        self._thread = QThread()
        self._worker = TranslationWorker(
            file_sheets=dict(self._file_sheets),
            pause_event=self._pause_event,
            cancel_event=self._cancel_event,
        )
        self._worker.moveToThread(self._thread)

        self._thread.started.connect(self._worker.run)
        self._worker.log.connect(self.append_log)
        self._worker.progress.connect(self._on_progress)
        self._worker.status.connect(lambda s: None)
        self._worker.file_started.connect(
            lambda name: self.lbl_current_file.setText(name)
        )
        self._worker.result_info.connect(self._store_result)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(lambda e: self.append_log("ERROR", f"Error: {e}"))

        self._thread.start()

    def _store_result(self, result: dict):
        self._last_result = result

    def _on_progress(self, current: int, total: int):
        if total > 0:
            self.bar_progress.setMaximum(total)
            self.bar_progress.setValue(current)
            percent = (current / total) * 100
            self.lbl_status.setText(f"Đang dịch... {current}/{total} cells ({percent:.1f}%)")

    def _on_finished(self, output_path: str):
        self.btn_translate.setEnabled(True)
        self.btn_pause.setEnabled(False)
        self.btn_resume.setEnabled(False)
        self.btn_cancel.setEnabled(False)
        self.btn_retry.setEnabled(bool(output_path))
        self.lbl_current_file.setText("")

        if self._thread:
            self._thread.quit()
            self._thread.wait()
            self._thread = None
            self._worker = None

        r = self._last_result
        status = r.get("status", "FAILED") if r else "FAILED"
        translated = r.get("translated", 0)
        failed = r.get("failed", 0)
        api_calls = r.get("api_calls", 0)
        cache_hits = r.get("cache_hits", 0)
        elapsed = r.get("elapsed", 0.0)
        out_paths = r.get("output_paths", [])
        out_str = "\n".join(out_paths) if out_paths else output_path

        # Summary log
        self.append_log(
            "INFO",
            f"[summary] translated={translated} failed={failed} "
            f"api_calls={api_calls} cache_hits={cache_hits} "
            f"elapsed={elapsed:.1f}s",
        )

        if status == "SUCCESS" and output_path:
            self.lbl_status.setText("Translation completed")
            self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px; color: #a6e3a1; font-weight: bold;")
            self.append_log("INFO", f"Done: {output_path}")
            report = f"Hoàn thành dịch thuật!\n- Tiến độ: 100%\n- Đã dịch: {translated} ô\n- Lỗi: {failed} ô\n- Thời gian hoàn thành: {elapsed:.1f} giây\n\nOutput:\n{out_str}"
            QMessageBox.information(
                self, "Báo cáo hoàn thành", report
            )
            if settings.get("processing_options.auto_open_file", False) and output_path:
                import os
                os.startfile(output_path)

        elif status == "PARTIAL":
            self.lbl_status.setText(f"Partial: {translated} ok, {failed} failed")
            self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px; color: orange; font-weight: bold;")
            QMessageBox.warning(
                self, "Partial",
                f"{translated} translated, {failed} failed.\n\nSee _LOG sheet.\n{out_str}",
            )

        elif status == "CANCELLED":
            self.lbl_status.setText(f"Cancelled ({translated} translated)")
            self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px; color: gray; font-weight: bold;")
            QMessageBox.warning(
                self, "Cancelled",
                f"Cancelled. {translated} cells saved.\n{out_str}",
            )

        else:
            err = r.get("error", "Unknown") if r else "Unknown"
            self.lbl_status.setText(f"Failed: {err[:60]}")
            self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px; color: red; font-weight: bold;")
            QMessageBox.critical(self, "Failed", err)

        self._update_translate_button()

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Retry Failed
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    def _on_retry_failed(self):
        r = self._last_result
        if not r:
            return
        out_paths = r.get("output_paths", [])
        if not out_paths:
            return

        import openpyxl
        from app.core.translators.batch_translator import CellItem, GeminiBatchTranslator
        from app.core.key_provider import KeyProvider

        total_retried = 0
        for out_path in out_paths:
            p = Path(out_path)
            if not p.exists():
                continue

            wb = openpyxl.load_workbook(str(p))
            if "_LOG" not in wb.sheetnames:
                wb.close()
                continue

            ws_log = wb["_LOG"]
            failed: list[tuple[str, str, str]] = []
            for row in ws_log.iter_rows(min_row=2, values_only=False):
                vals = [c.value for c in row]
                if len(vals) >= 5 and vals[4] == "FAILED" and vals[2]:
                    failed.append((vals[0] or "", vals[1] or "", vals[2] or ""))

            if not failed:
                wb.close()
                continue

            self.append_log("INFO", f"Retrying {len(failed)} failed cells in {p.name}")

            items = [CellItem(cell_id=addr, original=src) for _, addr, src in failed]

            kp = KeyProvider()
            bt = GeminiBatchTranslator(
                key_provider=kp,
                model_name=settings.get("selected_models.gemini", "gemini-3.1-flash-lite"),
                source_lang=settings.get("language_settings.source_lang", "Chinese"),
                target_lang=settings.get("language_settings.target_lang", "Vietnamese"),
                log_fn=lambda lvl, msg: self.append_log(lvl, msg),
            )
            result = bt.translate_batch(items)

            ok = 0
            for item, (sheet, addr, _) in zip(items, failed):
                if item.translated and sheet in wb.sheetnames:
                    ws = wb[sheet]
                    cell_ref = addr.split("!")[-1] if "!" in addr else addr
                    try:
                        ws[cell_ref].value = item.translated
                        ok += 1
                    except Exception:
                        pass

            if ok > 0:
                wb.save(str(p))
                self.append_log("INFO", f"Retried: {ok}/{len(failed)} fixed in {p.name}")
                total_retried += ok
            wb.close()

        if total_retried > 0:
            QMessageBox.information(self, "Retry Done", f"Fixed {total_retried} cells.")
        else:
            QMessageBox.information(self, "Retry", "No cells to retry or all retries failed.")

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Logging
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    _LOG_ICONS = {"INFO": "ℹ️", "WARN": "⚠️", "WARNING": "⚠️", "ERROR": "❌", "DEBUG": "🐛", "SUCCESS": "✅"}

    def append_log(self, level, msg):
        colors = {"INFO": "#89b4fa", "WARN": "#fab387", "ERROR": "#f38ba8", "DEBUG": "#a6adc8", "SUCCESS": "#a6e3a1"}
        color = colors.get(level, "#cdd6f4")
        icon = self._LOG_ICONS.get(level, "📋")
        self.txt_log.append(f'<span style="color:{color}">[{level}] {icon} {msg}</span>')

        if "Quota exceeded" in msg:
            self.lbl_status.setText("Waiting for API quota...")
            self.lbl_status.setStyleSheet("font-size: 13px; padding: 2px 4px; color: orange; font-weight: bold;")


