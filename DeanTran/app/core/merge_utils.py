"""
Merge utilities – merge multiple Excel files into a Master BOM workbook,
and merge multiple PDF files into a single PDF.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Callable, Optional

logger = logging.getLogger("DeanTran.merge_utils")

LogFn = Optional[Callable[[str, str], None]]


def _default_log(level: str, msg: str):
    logger.info("[%s] %s", level, msg)


def _sanitize_sheet_name(name: str) -> str:
    """
    Sanitize a sheet name to make it fully compliant with Excel requirements:
    1. Strip leading and trailing spaces and single quotes.
    2. Remove forbidden characters: \ / ? * [ ] :
    3. Ensure it is not empty (fallback to "Sheet").
    4. Truncate to maximum 31 characters.
    5. Re-strip space and quote in case truncation left a trailing space or quote.
    """
    if not name:
        return "Sheet"
    for char in ['\\', '/', '?', '*', '[', ']', ':']:
        name = name.replace(char, ' ')
    name = name.strip().strip("'").strip()
    name = name[:31]
    name = name.strip().strip("'").strip()
    if not name:
        return "Sheet"
    return name


# ─── Excel Merge ────────────────────────────────────────────────────────

def merge_excel_to_master(
    file_paths: list[str | Path],
    output_path: str | Path,
    log_fn: LogFn = None,
) -> Path:
    """
    Merge multiple .xlsx files into a single Master BOM workbook.
    Primary method: Excel COM automation (100% formatting preservation
    including images, charts, conditional formatting, etc.).
    Fallback: openpyxl cell-by-cell copy (if Excel is not installed).
    """
    raw_log = log_fn or _default_log

    # Safe log wrapper to prevent encoding/charmap crashes on Windows console
    def _log(level: str, msg: str):
        try:
            raw_log(level, msg)
        except Exception:
            try:
                safe_msg = msg.encode('ascii', errors='replace').decode('ascii')
                raw_log(level, safe_msg)
            except Exception:
                pass

    try:
        result = _merge_excel_com(file_paths, output_path, _log)
        return result
    except Exception as exc:
        _log("WARN", f"COM merge failed ({exc}), falling back to openpyxl...")
        return _merge_excel_openpyxl(file_paths, output_path, _log)


def _merge_excel_com(
    file_paths: list[str | Path],
    output_path: str | Path,
    _log: Callable,
) -> Path:
    """
    Merge Excel files using COM automation.
    Requires Microsoft Excel installed on the machine.
    Preserves 100% formatting: images, charts, conditional formatting,
    data validation, hyperlinks, comments, rich text, auto-filters, etc.
    """
    import win32com.client
    import pythoncom

    output_path = Path(output_path).resolve()
    # Resolve all paths to absolute
    abs_paths = [str(Path(fp).resolve()) for fp in file_paths]

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # Create the master workbook
        master_wb = excel.Workbooks.Add()
        used_names: set[str] = set()
        total_sheets = 0

        for file_idx, fp in enumerate(abs_paths, 1):
            fp_name = Path(fp).name
            fp_stem = Path(fp).stem
            _log("INFO", f"[{file_idx}/{len(abs_paths)}] Reading: {fp_name}")
            try:
                src_wb = excel.Workbooks.Open(fp, ReadOnly=True)
            except Exception as exc:
                _log("ERROR", f"Cannot open {fp_name}: {exc}")
                continue

            for i in range(1, src_wb.Sheets.Count + 1):
                src_sheet = src_wb.Sheets(i)
                base_name = src_sheet.Name
                sheet_name = _sanitize_sheet_name(base_name)

                # Resolve name conflicts
                if sheet_name in used_names:
                    sheet_name = _sanitize_sheet_name(f"{fp_stem}_{base_name}")
                counter = 2
                original = sheet_name
                while sheet_name in used_names:
                    suffix = f"_{counter}"
                    sheet_name = _sanitize_sheet_name(original[:31 - len(suffix)] + suffix)
                    counter += 1
                used_names.add(sheet_name)

                _log("INFO", f"  Sheet: {base_name} -> {sheet_name}")
                # Copy entire sheet to master workbook (preserves everything)
                src_sheet.Copy(After=master_wb.Sheets(master_wb.Sheets.Count))
                copied_sheet = master_wb.Sheets(master_wb.Sheets.Count)
                copied_sheet.Name = sheet_name
                total_sheets += 1

            src_wb.Close(SaveChanges=False)

        # Remove default empty sheets created with the new workbook
        while master_wb.Sheets.Count > total_sheets:
            try:
                master_wb.Sheets(1).Delete()
            except Exception:
                break

        # Save as xlsx (FileFormat=51 = xlOpenXMLWorkbook)
        master_wb.SaveAs(str(output_path), FileFormat=51)
        master_wb.Close(SaveChanges=False)
        _log("INFO", f"✅ Master BOM saved: {output_path.name} ({total_sheets} sheets)")

    except Exception as exc:
        _log("ERROR", f"Excel COM merge failed: {exc}")
        raise
    finally:
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    return output_path


def _merge_excel_openpyxl(
    file_paths: list[str | Path],
    output_path: str | Path,
    _log: Callable,
) -> Path:
    """
    Fallback: Merge Excel files using openpyxl (cell-by-cell copy).
    Preserves: values, font, border, fill, number_format, protection,
    alignment, merged cells, column widths, row heights.
    Does NOT preserve: images, charts, conditional formatting, etc.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    from copy import copy

    output_path = Path(output_path)

    master = openpyxl.Workbook()
    # Remove the default sheet
    default_sheet = master.active
    if default_sheet is not None:
        master.remove(default_sheet)

    used_names: set[str] = set()

    for file_idx, fp in enumerate(file_paths, 1):
        fp = Path(fp)
        _log("INFO", f"[{file_idx}/{len(file_paths)}] Reading: {fp.name}")

        try:
            wb = openpyxl.load_workbook(str(fp))
        except Exception as exc:
            _log("ERROR", f"Cannot open {fp.name}: {exc}")
            continue

        file_stem = fp.stem

        for ws in wb.worksheets:
            base_name = ws.title
            sheet_name = _sanitize_sheet_name(base_name)

            # If name conflict, prefix with file stem
            if sheet_name in used_names:
                sheet_name = _sanitize_sheet_name(f"{file_stem}_{base_name}")

            counter = 2
            original = sheet_name
            while sheet_name in used_names:
                suffix = f"_{counter}"
                sheet_name = _sanitize_sheet_name(original[: 31 - len(suffix)] + suffix)
                counter += 1

            used_names.add(sheet_name)
            new_ws = master.create_sheet(title=sheet_name)
            _log("INFO", f"  Sheet: {ws.title} -> {sheet_name}")

            # Copy column widths
            for col_letter, col_dim in ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = col_dim.width
                new_ws.column_dimensions[col_letter].hidden = col_dim.hidden

            # Copy row heights
            for row_num, row_dim in ws.row_dimensions.items():
                new_ws.row_dimensions[row_num].height = row_dim.height
                new_ws.row_dimensions[row_num].hidden = row_dim.hidden

            # Copy cells
            for row in ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(
                        row=cell.row,
                        column=cell.column,
                        value=cell.value,
                    )
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            # Copy merged cells
            for merged_range in ws.merged_cells.ranges:
                new_ws.merge_cells(str(merged_range))

        wb.close()

    master.save(str(output_path))
    master.close()
    _log("INFO", f"✅ Master BOM saved: {output_path.name} ({len(used_names)} sheets)")
    return output_path


# ─── PDF Merge ──────────────────────────────────────────────────────────

def merge_pdfs(
    pdf_paths: list[str | Path],
    output_path: str | Path,
    log_fn: LogFn = None,
) -> Path:
    """
    Merge multiple PDF files into a single PDF using pypdfium2.
    """
    import pypdfium2 as pdfium

    _log = log_fn or _default_log
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    dest_pdf = pdfium.PdfDocument.new()
    total_pages = 0

    for file_idx, fp in enumerate(pdf_paths, 1):
        fp = Path(fp)
        _log("INFO", f"[{file_idx}/{len(pdf_paths)}] Merging: {fp.name}")

        try:
            src_pdf = pdfium.PdfDocument(str(fp))
            n_pages = len(src_pdf)
            dest_pdf.import_pages(src_pdf)
            total_pages += n_pages
            _log("INFO", f"  Added {n_pages} page(s)")
            src_pdf.close()
        except Exception as exc:
            _log("ERROR", f"Cannot merge {fp.name}: {exc}")
            continue

    dest_pdf.save(str(output_path))
    dest_pdf.close()

    _log("INFO", f"✅ Merged PDF saved: {output_path.name} ({total_pages} pages)")
    return output_path


# ─── Word Merge ─────────────────────────────────────────────────────────

def merge_word_documents(
    doc_paths: list[str | Path],
    output_path: str | Path,
    log_fn: LogFn = None,
) -> Path:
    """
    Merge multiple .docx files into a single Word document using COM automation.

    Uses Microsoft Word directly via win32com to insert each sub-document,
    guaranteeing 100% content preservation (no page loss).
    """
    import win32com.client
    import pythoncom

    _log = log_fn or _default_log
    output_path = Path(output_path).resolve()

    if not doc_paths:
        raise ValueError("No documents to merge")

    # Resolve all paths to absolute
    abs_paths = [str(Path(fp).resolve()) for fp in doc_paths]

    pythoncom.CoInitialize()
    word = None
    try:
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        word.DisplayAlerts = False

        # Open the first document
        _log("INFO", f"[1/{len(abs_paths)}] Base document: {Path(abs_paths[0]).name}")
        doc = word.Documents.Open(abs_paths[0])

        for file_idx, fp in enumerate(abs_paths[1:], 2):
            _log("INFO", f"[{file_idx}/{len(abs_paths)}] Appending: {Path(fp).name}")
            try:
                # Move cursor to end of document
                rng = doc.Content
                rng.Collapse(0)  # wdCollapseEnd = 0

                # Insert a page break
                rng.InsertBreak(7)  # wdPageBreak = 7

                # Move cursor to very end again
                rng = doc.Content
                rng.Collapse(0)

                # Insert the sub-document file
                rng.InsertFile(fp)
                _log("INFO", f"  ✓ Inserted {Path(fp).name}")

            except Exception as exc:
                _log("ERROR", f"Cannot insert {Path(fp).name}: {exc}")
                continue

        # Save as the output file
        doc.SaveAs2(str(output_path), 16)  # wdFormatDocumentDefault = 16 (.docx)
        doc.Close(False)

        _log("INFO", f"✅ Merged Word saved: {output_path.name} ({len(abs_paths)} files)")

    except Exception as exc:
        _log("ERROR", f"Word merge failed: {exc}")
        raise
    finally:
        if word:
            word.Quit(False)
        pythoncom.CoUninitialize()

    return output_path


# ─── PowerPoint Merge ───────────────────────────────────────────────────

def merge_pptx(
    pptx_paths: list[str | Path],
    output_path: str | Path,
    log_fn: LogFn = None,
) -> Path:
    """
    Merge multiple .pptx files into a single presentation.
    Uses PowerPoint COM automation to guarantee 100% fidelity (themes, master layouts, media, animations).
    """
    import win32com.client
    import pythoncom
    import shutil

    _log = log_fn or _default_log
    output_path = Path(output_path).resolve()
    
    if not pptx_paths:
        raise ValueError("No presentations to merge")
        
    abs_paths = [str(Path(fp).resolve()) for fp in pptx_paths]
    
    # Copy the first presentation as the base
    _log("INFO", f"[1/{len(abs_paths)}] Base presentation: {Path(abs_paths[0]).name}")
    shutil.copy2(abs_paths[0], str(output_path))
    
    if len(abs_paths) == 1:
        return output_path

    pythoncom.CoInitialize()
    ppt_app = None
    try:
        ppt_app = win32com.client.Dispatch("PowerPoint.Application")
        # Try WithWindow=False first, fallback to standard Open if WPS complains
        try:
            prs = ppt_app.Presentations.Open(str(output_path), WithWindow=False)
        except Exception:
            prs = ppt_app.Presentations.Open(str(output_path))
        
        for file_idx, fp in enumerate(abs_paths[1:], 2):
            _log("INFO", f"[{file_idx}/{len(abs_paths)}] Appending: {Path(fp).name}")
            try:
                slide_count = prs.Slides.Count
                prs.Slides.InsertFromFile(fp, slide_count)
                _log("INFO", f"  ✓ Inserted slides from {Path(fp).name}")
            except Exception as exc:
                _log("ERROR", f"Cannot insert slides from {Path(fp).name}: {exc}")
                continue
                
        prs.Save()
        prs.Close()
        _log("INFO", f"✅ Merged PPT saved: {output_path.name}")
    except Exception as exc:
        _log("ERROR", f"PPT merge failed: {exc}")
        raise
    finally:
        if ppt_app:
            try:
                ppt_app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
        
    return output_path


# ─── Merge filename helper ──────────────────────────────────────────────

def build_merge_filename(
    file_paths: list[str | Path],
    extension: str,
) -> str:
    """
    Build a merge output filename from the first file's stem.
    Example: ['BOM_1.xlsx', 'BOM_2.xlsx'] → 'BOM_1_merge.xlsx'
    """
    first_stem = Path(file_paths[0]).stem
    ext = extension if extension.startswith(".") else f".{extension}"
    return f"{first_stem}_merge{ext}"
