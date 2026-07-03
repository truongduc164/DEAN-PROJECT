"""
ExcelProcessor – settings-driven Excel translation pipeline with BATCH support.

Output naming: Book1.xlsx → Book1_Vi.xlsx (same directory as input)
Sheet duplication: for each "SheetX", create "SheetX_Vi" with blue tab color.
Original sheets are kept unchanged.  Images are preserved.
"""
from __future__ import annotations

import copy
import re
import threading
import time
import traceback
from copy import deepcopy
from datetime import datetime, date as dt_date, time as dt_time
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

try:
    from PIL import JpegImagePlugin
    # Prevent Pillow from treating JPEG images as MPO containers,
    # which causes openpyxl to encounter unknown '.mpo' extension.
    JpegImagePlugin._getmp = lambda self: None
except ImportError:
    pass

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# Fix KeyError('.mpo'): openpyxl uses its own MimeTypes instance in
# openpyxl.packaging.manifest.  Register .mpo there so wb.save() works
# when the workbook contains MPO images.
try:
    from openpyxl.packaging import manifest as _manifest
    _manifest.mimetypes.add_type('image/jpeg', '.mpo')
except Exception:
    pass

from app.core.event_manager import EventManager
from app.core.output_naming import (
    get_unique_output_path as _get_unique_path,
    build_output_path as _build_base_path,
    get_lang_suffix,
    LANG_SUFFIX as _LANG_SUFFIX,
)
from app.core.translators.translator_service import BaseTranslator
from app.core.pinyin_helper import (
    get_pinyin_line,
    is_available as pinyin_available,
    render_text_with_pinyin,
)
from app.term_engine.glossary_loader import GlossaryLoader
from app.term_engine.term_override import TermOverride
from app.term_engine import regex_protection

# Excel standard indexed color palette (0-63).
# Maps index number → RRGGBB hex string.
_INDEXED_COLORS = [
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",  # 0-7
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",  # 8-15
    "800000", "008000", "000080", "808000", "800080", "008080", "C0C0C0", "808080",  # 16-23
    "9999FF", "993366", "FFFFCC", "CCFFFF", "660066", "FF8080", "0066CC", "CCCCFF",  # 24-31
    "000080", "FF00FF", "FFFF00", "00FFFF", "800080", "800000", "008080", "0000FF",  # 32-39
    "00CCFF", "CCFFFF", "CCFFCC", "FFFF99", "99CCFF", "FF99CC", "CC99FF", "FFCC99",  # 40-47
    "3366FF", "33CCCC", "99CC00", "FFCC00", "FF9900", "FF6600", "666699", "969696",  # 48-55
    "003366", "339966", "003300", "333300", "993300", "993366", "333399", "333333",  # 56-63
]


def _indexed_to_argb(idx_str: str) -> str:
    """Convert an indexed color attribute to FFRRGGBB."""
    try:
        idx = int(idx_str)
        if 0 <= idx < len(_INDEXED_COLORS):
            return f"FF{_INDEXED_COLORS[idx]}"
    except (ValueError, TypeError):
        pass
    return "FF000000"


def _parse_rPr(rPr_el, NS: str, theme_colors: list[str] = None) -> dict:
    """Extract font info from an XML <rPr> element."""
    info: dict = {"name": "Calibri", "sz": 11, "color": "FF000000", "b": False, "i": False}
    rFont = rPr_el.find(f"{{{NS}}}rFont")
    if rFont is not None:
        info["name"] = rFont.get("val", "Calibri")
    sz = rPr_el.find(f"{{{NS}}}sz")
    if sz is not None:
        try:
            info["sz"] = float(sz.get("val", "11"))
        except ValueError:
            pass
    color = rPr_el.find(f"{{{NS}}}color")
    if color is not None:
        rgb = color.get("rgb")
        if rgb and rgb != "00000000":
            info["color"] = rgb if len(rgb) == 8 else f"FF{rgb}"
        else:
            theme = color.get("theme")
            indexed = color.get("indexed")
            if theme is not None and theme_colors:
                try:
                    ti = int(theme)
                    if ti < len(theme_colors):
                        info["color"] = f"FF{theme_colors[ti]}"
                except ValueError:
                    pass
            elif indexed is not None:
                info["color"] = _indexed_to_argb(indexed)
    info["b"] = rPr_el.find(f"{{{NS}}}b") is not None
    info["i"] = rPr_el.find(f"{{{NS}}}i") is not None
    return info



# ── Output path helper (delegates to shared module) ──────────────────

def build_output_path(input_path: Path, target_lang: str) -> Path:
    """Build BASE output path (backward-compat public API).

    Book1.xlsx + target_lang="Vietnamese" → Book1_Vi.xlsx
    """
    return _build_base_path(input_path, target_lang)


# ── Sheet name helpers ───────────────────────────────────────────────

def _make_translated_sheet_name(
    original_name: str, target_lang: str, existing_names: set[str],
) -> str:
    """
    Create translated sheet name, respecting Excel 31-char limit.
    E.g. "Sheet1" → "Sheet1_Vi". Handles collisions with _2, _3, etc.
    """
    suffix_code = get_lang_suffix(target_lang)
    suffix = f"_{suffix_code}"

    max_base = 31 - len(suffix)
    base = original_name[:max_base]
    candidate = f"{base}{suffix}"

    # Collision handling
    counter = 2
    while candidate in existing_names:
        num_suffix = f"_{counter}"
        max_base2 = 31 - len(suffix) - len(num_suffix)
        base2 = original_name[:max_base2]
        candidate = f"{base2}{num_suffix}{suffix}"
        counter += 1

    return candidate


BLUE_TAB_COLOR = "0000FF"
_DATETIME_TEXT_PATTERNS = (
    re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?$"),
    re.compile(r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?$"),
    re.compile(r"^\d{1,2}:\d{2}(?::\d{2})?$"),
)


class ExcelProcessor:
    """
    Settings-driven Excel translation pipeline with batch support.

    Output naming: ``input.xlsx`` → ``input_Vi.xlsx`` (same directory)
    Sheet duplication: creates translated copy sheets with blue tab color.
    """

    def __init__(
        self,
        translator: BaseTranslator,
        event_manager: Optional[EventManager] = None,
        skip_empty: bool = True,
        skip_numeric: bool = True,
        source_lang: str = "Chinese",
        target_lang: str = "Vietnamese",
        glossary: Optional[GlossaryLoader] = None,
        use_glossary: bool = False,
        use_protection: bool = True,
        prompt: str = "",
        max_rows: int = 8000,
        max_cols: int = 50,
        min_interval: float = 0.2,
        batch_mode: bool = True,
        sheets_to_translate: Optional[List[str]] = None,
        pause_event: Optional[threading.Event] = None,
        cancel_event: Optional[threading.Event] = None,
        output_mode: str = "new_sheet",
        sheet_target_rows: Optional[Dict[str, int]] = None,
    ) -> None:
        self.translator = translator
        self.em = event_manager or EventManager()
        self.skip_empty = skip_empty
        self.skip_numeric = skip_numeric
        self.source_lang = source_lang
        self.target_lang = target_lang
        self.prompt = prompt
        self.use_protection = use_protection
        self.max_rows = max_rows
        self.max_cols = max_cols
        self.min_interval = min_interval
        self.sheets_to_translate = sheets_to_translate  # None = all sheets
        self.output_mode = output_mode  # 'overwrite'|'prefix'|'suffix'|'new_sheet'
        self.sheet_target_rows = sheet_target_rows or {}  # {sheet_name: start_row}

        # Threading events for Pause / Cancel
        self._pause_event = pause_event or threading.Event()
        if pause_event is None:
            self._pause_event.set()  # default: running (not paused)
        self._cancel_event = cancel_event or threading.Event()

        self.batch_mode = batch_mode

        # Glossary
        self.use_glossary = use_glossary and glossary is not None and not glossary.is_empty()
        self._term_override: Optional[TermOverride] = None
        if self.use_glossary and glossary is not None:
            self._term_override = TermOverride(glossary)

        # Track results for status
        self._translated_count: int = 0
        self._error_count: int = 0
        self._api_calls: int = 0
        self._cache_hits: int = 0
        self._error_log: List[Tuple[str, str, str, str, str]] = []
        #  (sheet, cell_addr, source, translated_text, error_msg)
        self._rich_text_records: list = []  # for XLSX post-processing

        # Initialize OCR pipeline for image translation
        from app.core.ocr.image_translation_pipeline import ImageTranslationPipeline
        self.ocr_pipeline = ImageTranslationPipeline(self.translator, self.em)

        # Translated text format mode (shared with PPT/Word UI settings)
        from app.settings.settings_manager import settings
        translated_mode = settings.get("text_style_settings.translated_text_format_mode", None)
        keep_format_legacy = settings.get("text_style_settings.keep_format", True)
        if translated_mode in ("keep_original_format", "custom_format"):
            self._text_format_mode = translated_mode
        else:
            self._text_format_mode = "keep_original_format" if keep_format_legacy else "custom_format"

        self._custom_font_name = settings.get("text_style_settings.font_family", "Arial")
        self._custom_font_size = settings.get("text_style_settings.font_size", 14)
        self._custom_font_color = settings.get("text_style_settings.font_color", "#000000")
        self._custom_bold = settings.get("text_style_settings.bold", False)
        self._custom_italic = settings.get("text_style_settings.italic", False)
        self._custom_underline = settings.get("text_style_settings.underline", False)
        self._add_chinese_pinyin = settings.get("processing_options.add_chinese_pinyin", False)
        self._pinyin_font_size = settings.get("processing_options.pinyin_font_size", 10)
        self._pinyin_font_family = settings.get("processing_options.pinyin_font_family", "Arial")
        self._pinyin_font_color = settings.get("processing_options.pinyin_font_color", "#888888")
        self._pinyin_format_mode = settings.get("processing_options.pinyin_format_mode", "custom")
        if self._add_chinese_pinyin and not pinyin_available():
            self.em.log("WARN", "Pinyin is enabled but pypinyin is not installed; skipping pinyin.")

    def _pinyin_for_lang(self, text: str, lang: str) -> str:
        if not self._add_chinese_pinyin or lang != "Chinese":
            return ""
        return get_pinyin_line(text)

    def _render_with_optional_pinyin(self, text: str, lang: str) -> str:
        if not text:
            return ""
        if not self._add_chinese_pinyin or lang != "Chinese":
            return text
        return render_text_with_pinyin(text)

    def _pinyin_inline_font(self, base_font):
        """Create an InlineFont for pinyin text."""
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles import Font

        base = base_font or Font()
        if self._pinyin_format_mode == "keep_original":
            return self._to_inline_font(base)
        return InlineFont(
            rFont=self._pinyin_font_family or base.name,
            charset=base.charset,
            family=base.family,
            b=False,
            i=True,
            strike=base.strike,
            outline=base.outline,
            shadow=base.shadow,
            condense=base.condense,
            extend=base.extend,
            color=self._hex_to_argb(self._pinyin_font_color),
            sz=self._pinyin_font_size,
            u=None,
            vertAlign=base.vertAlign,
            scheme=base.scheme,
        )

    def _segment_rich_text(self, text: str, lang: str, inline_font, cell_font=None):
        from openpyxl.cell.rich_text import TextBlock

        if not text:
            return []

        pinyin_font = self._pinyin_inline_font(cell_font) if self._add_chinese_pinyin else inline_font
        blocks: list[TextBlock] = []
        raw_lines = text.replace("\r", "").split("\n")
        for idx, raw_line in enumerate(raw_lines):
            if idx > 0:
                blocks.append(TextBlock(inline_font, "\n"))
            if raw_line:
                blocks.append(TextBlock(inline_font, raw_line))

            pinyin = self._pinyin_for_lang(raw_line, lang)
            if pinyin:
                blocks.append(TextBlock(pinyin_font, "\n"))
                blocks.append(TextBlock(pinyin_font, pinyin))
        return blocks

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def process(self, file_path: str | Path, output_dir: Optional[str | Path] = None) -> Path:
        file_path = Path(file_path)
        self.em.log("INFO", f"Loading workbook: {file_path.name}")

        from app.settings.settings_manager import settings

        # Pre-read original cell formatting from raw XLSX XML
        self._original_cell_fonts = self._preload_cell_fonts(file_path)

        # Build output path
        out_path = _get_unique_path(file_path, self.target_lang, output_dir=output_dir)
        self.em.log("INFO", f"input_path={file_path}")
        self.em.log("INFO", f"output_path={out_path}")

        # Log settings
        self.em.log("INFO", f"Max rows: {self.max_rows}, Max cols: {self.max_cols}")
        self.em.log("INFO", f"Min interval: {self.min_interval}s")
        self.em.log("INFO", f"Batch mode: {self.batch_mode}")

        # Pre-determine original sheets to scan
        # We need to know this before duplication to know what to copy
        import openpyxl
        temp_wb = openpyxl.load_workbook(str(file_path), read_only=True)
        original_sheets = list(temp_wb.sheetnames)
        temp_wb.close()

        if self.sheets_to_translate:
            sheets_todo = [s for s in original_sheets if s in self.sheets_to_translate]
        else:
            sheets_todo = original_sheets

        if not sheets_todo:
            self.em.log("WARN", "No sheets selected for translation!")
            wb = openpyxl.load_workbook(str(file_path))
            wb.save(str(out_path))
            return out_path

        self.em.log("INFO", f"sheets_to_translate={sheets_todo}")
        self.em.log("INFO", f"output_mode={self.output_mode}")

        # Duplicate sheets via COM to preserve charts and images
        temp_file = None
        sheet_mapping: dict[str, str] = {}
        try:
            self.em.log("INFO", "Duplicating sheets via COM to preserve charts and images...")
            temp_file, sheet_mapping = self._duplicate_sheets_via_com(file_path, sheets_todo, self.target_lang)
            wb: Workbook = openpyxl.load_workbook(str(temp_file))
            self.em.log("INFO", "COM sheet duplication successful.")
        except Exception as exc:
            self.em.log("WARN", f"COM sheet duplication failed: {exc}. Falling back to openpyxl sheet copy.")
            temp_file = None
            wb: Workbook = openpyxl.load_workbook(str(file_path))

        # Always create translated sheet copies if not done via COM
        if temp_file is not None:
            target_sheets = list(sheet_mapping.values())
        else:
            sheet_mapping = {}
            existing_names = set(wb.sheetnames)

            for orig_name in sheets_todo:
                trans_name = _make_translated_sheet_name(
                    orig_name, self.target_lang, existing_names,
                )
                existing_names.add(trans_name)
                sheet_mapping[orig_name] = trans_name

                # Copy sheet
                orig_ws = wb[orig_name]
                trans_ws = wb.copy_worksheet(orig_ws)
                trans_ws.title = trans_name
                trans_ws.sheet_properties.tabColor = BLUE_TAB_COLOR

                # Preserve images (openpyxl.copy_worksheet does NOT copy images)
                self._copy_images(orig_ws, trans_ws)

                # Move translated sheet to sit right after the original
                orig_idx = wb.sheetnames.index(orig_name)
                trans_idx = wb.sheetnames.index(trans_name)
                offset = (orig_idx + 1) - trans_idx
                if offset != 0:
                    wb.move_sheet(trans_ws, offset=offset)

            target_sheets = list(sheet_mapping.values())
            
        self.em.log("INFO", f"created_sheets={target_sheets}")
        self.temp_file_path = temp_file  # Save path to delete later


        # Scan translatable cells from the TARGET sheets
        cells_map = self._scan_cells(wb, sheet_names=target_sheets)
        total = len(cells_map)
        self.em.log("INFO", f"total_cells={self._count_all_cells(wb, target_sheets)}")
        self.em.log("INFO", f"translatable_cells={total}")

        # Collect images for OCR (if enabled) â€“ scan ORIGINAL sheets (not translated copies)
        images_to_ocr: list = []
        if settings.get("ocr_settings.image_text_translation_enabled", False):
            # Use target_sheets first (they should have images via _copy_images),
            # but fall back to original sheets if target sheets are empty
            images_to_ocr = self._collect_xlsx_images(wb, target_sheets)
            if not images_to_ocr:
                # No images on translated copies â€“ try original sheets
                images_to_ocr = self._collect_xlsx_images(wb, sheets_todo)
                if images_to_ocr:
                    self.em.log("INFO", f"Collected {len(images_to_ocr)} images from original sheets (fallback)")
        total_images = len(images_to_ocr)
        if total_images > 0:
            self.em.log("INFO", f"images_found={total_images}")

        if total == 0 and total_images == 0:
            self.em.log("WARN", "No translatable cells or images found!")
            wb.save(str(out_path))
            self.em.status("Done (nothing to translate)")
            return out_path

        if self.use_glossary:
            self.em.log("INFO", "Glossary override: ENABLED")
        if self.use_protection:
            self.em.log("INFO", "Regex protection: ENABLED")
        if self.prompt:
            self.em.log("INFO", f"Prompt: {self.prompt[:80]}…")

        self.em.progress(0, total)

        # Reset counters
        self._translated_count = 0
        self._error_count = 0
        self._api_calls = 0
        self._cache_hits = 0
        self._error_log = []

        t0 = time.time()

        if self.batch_mode:
            self._process_batch_mode(wb, cells_map, total)
        else:
            from app.core.translators.translator_service import MockTranslator
            if not isinstance(self.translator, MockTranslator):
                self.em.log("ERROR", "Cell-by-cell mode is disabled to protect API quota! Batch mode required.")
                self.em.log("ERROR", "Translation aborted.")
                self._error_count = total
                self._cancel_event.set()
            else:
                self._process_cell_by_cell(wb, cells_map, total)

        elapsed = time.time() - t0
        cells_per_sec = self._translated_count / elapsed if elapsed > 0 else 0

        # Determine result status
        cancelled = self._cancel_event.is_set()
        if cancelled or self._error_count >= (total // 2 + 1): # Circuit breaker also triggers save
            result_status = "CANCELLED" if cancelled else "FAILED_CIRCUIT"
            # Rename output to _stop suffix
            stop_name = file_path.stem + "_stop" + file_path.suffix
            out_path = file_path.parent / stop_name
            self.em.log("WARN", f"{result_status} → saving partial result as {stop_name}")
            
            # Save checkpoint state into _stop.json
            import json
            stop_json_path = out_path.with_suffix(".json")
            checkpoint = {
                "file": str(file_path),
                "total": total,
                "translated": self._translated_count,
                "failed": self._error_count,
                "error_log": self._error_log,
                "last_run": datetime.now().isoformat()
            }
            try:
                stop_json_path.write_text(json.dumps(checkpoint, ensure_ascii=False, indent=2), encoding="utf-8")
                self.em.log("INFO", f"Checkpoint saved to {stop_json_path.name}")
            except Exception as e:
                self.em.log("WARN", f"Failed to save checkpoint: {e}")
                
        elif self._translated_count == 0 and self._error_count > 0:
            result_status = "FAILED"
        elif self._error_count > 0:
            result_status = "PARTIAL"
        else:
            result_status = "SUCCESS"

        # ── Translate images (OCR) ───────────────────────────
        if total_images > 0:
            self._process_xlsx_images(images_to_ocr, wb)

        # Save output
        out_path = self._safe_save(wb, out_path)

        # Post-process: convert plain shared strings to rich text shared strings
        if hasattr(self, '_rich_text_records') and self._rich_text_records:
            try:
                self._post_process_rich_text(out_path)
                self.em.log('INFO', f'Rich text applied to {len(self._rich_text_records)} cells')
            except Exception as exc:
                self.em.log('WARN', f'Rich text post-processing failed: {exc}')

        # Final status
        self.em.log(
            "INFO",
            f"result={result_status} translated={self._translated_count} "
            f"errors={self._error_count} api_calls={self._api_calls} "
            f"cache_hits={self._cache_hits} "
            f"elapsed={elapsed:.1f}s cells/sec={cells_per_sec:.1f} "
            f"output_path={out_path}",
        )

        if result_status == "FAILED":
            self.em.log("ERROR", "ZERO cells translated! Check logs for errors.")
            self.em.status("FAILED (Gemini error)")
        elif result_status == "CANCELLED":
            self.em.status(f"Cancelled ({self._translated_count} translated so far)")
        elif result_status == "PARTIAL":
            self.em.status(f"Done ({self._translated_count} translated, {self._error_count} errors)")
        else:
            self.em.status("Done")

        self.em.log("INFO", f"Saved: {out_path.name}")

        # Store result info for UI
        self.last_result = {
            "status": result_status,
            "output_path": str(out_path),
            "translated": self._translated_count,
            "failed": self._error_count,
            "api_calls": self._api_calls,
            "cache_hits": self._cache_hits,
            "elapsed": elapsed,
        }
        # Clean up temporary COM file
        if getattr(self, 'temp_file_path', None) is not None:
            try:
                self.temp_file_path.unlink()
                self.em.log("INFO", "Temporary COM duplication file cleaned up.")
            except Exception as e:
                self.em.log("WARN", f"Could not remove temp file: {e}")

        return out_path

    # ------------------------------------------------------------------
    # Batch mode (fast)
    # ------------------------------------------------------------------

    def _process_batch_mode(self, wb: Workbook, cells_map: dict, total: int):
        from app.core.key_provider import KeyProvider
        from app.core.translators.batch_translator import (
            GeminiBatchTranslator, CellItem,
        )
        from app.settings.settings_manager import settings

        model = settings.get("selected_models.gemini", "gemini-3.1-flash-lite")

        kp = KeyProvider()
        kp.log_status()

        # Check if DeepSeek key is available (for DeepSeek tool)
        _tool = settings.get("translation_tool", "Gemini")
        _has_deepseek_key = False
        if _tool == "DeepSeek":
            try:
                import os
                from app.core.secure_storage import SecureStorage
                _ds_key = SecureStorage().load_deepseek_key() or os.environ.get("DEEPSEEK_API_KEY", "")
                _has_deepseek_key = bool(_ds_key and _ds_key.strip())
            except Exception:
                pass
        _keys_available = kp.key_loaded or _has_deepseek_key

        # Build CellItem list
        items: list[CellItem] = []
        for cell_id, (sheet_name, row_idx, col_idx) in cells_map.items():
            ws = wb[sheet_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            text = str(cell.value)
            
            # Khởi tạo CellItem
            item = CellItem(cell_id=cell_id, original=text)
            item.real_original = text

            # Glossary pre-replace
            if self.use_glossary and getattr(self, '_term_override', None):
                term_dict = self._term_override.glossary.get_dict() if self._term_override.glossary else {}
                matched_val = None
                
                # Check if original was EXACTLY in dictionary (case-insensitive)
                lower_text = text.strip().lower()
                for k, v in term_dict.items():
                    if k.lower() == lower_text:
                        matched_val = v
                        break
                        
                replaced_text = self._term_override.pre_replace(text)
                
                if matched_val:
                    item.translated = matched_val
                elif replaced_text != text and not _keys_available:
                    item.translated = replaced_text
                
                item.original = replaced_text
            
            if not _keys_available and not item.translated:
                item.translated = item.original
                item.error = "No API key available"

            items.append(item)
            
        if not _keys_available:
            self.em.log("WARN", "No API key available. Only dictionary terms will be translated.")

        # Create batch translator
        bt = GeminiBatchTranslator(
            key_provider=kp,
            model_name=model,
            source_lang=self.source_lang,
            target_lang=self.target_lang,
            prompt=self.prompt,
            min_interval=self.min_interval,
            log_fn=lambda lvl, msg: self.em.log(lvl, msg),
            pause_event=self._pause_event,
            cancel_event=self._cancel_event,
            progress_fn=lambda cur, tot: self.em.progress(cur, tot),
        )

        # Run batch translation
        result = bt.translate_batch(items)

        self._api_calls = result.api_calls
        self._cache_hits = result.cache_hits

        self.em.log(
            "INFO",
            f"Batch complete: api_calls={result.api_calls} "
            f"cache_hits={result.cache_hits} errors={result.errors} "
            f"elapsed={result.elapsed:.1f}s",
        )

        # Write results back to workbook (translated sheet only)
        done = 0
        for item in items:
            cell_id = item.cell_id
            if cell_id not in cells_map:
                continue
            sheet_name, row_idx, col_idx = cells_map[cell_id]
            ws = wb[sheet_name]
            cell = ws.cell(row=row_idx, column=col_idx)

            if item.translated:
                translated = item.translated
                if self.use_glossary and getattr(self, '_term_override', None):
                    translated = self._term_override.post_fix(translated)
                real_orig = getattr(item, 'real_original', item.original)
                # Skip write if text unchanged (source-language filter)
                if translated.strip() != real_orig.strip():
                    self._write_cell_result(cell, real_orig, translated)
                self._translated_count += 1
            else:
                # Cell not translated — log error
                self._error_count += 1
                self._error_log.append((
                    sheet_name,
                    cell_id,
                    item.original,
                    "",
                    item.error or "Gemini returned no translation",
                ))

            done += 1
            self.em.progress(done, total)

        # Also capture batch-level errors
        if result.errors > 0 and not self._error_log:
            self._error_count += result.errors

    # ------------------------------------------------------------------
    # Cell-by-cell fallback (legacy)
    # ------------------------------------------------------------------

    def _process_cell_by_cell(self, wb: Workbook, cells_map: dict, total: int):
        done = 0
        for cell_id, (sheet_name, row_idx, col_idx) in cells_map.items():
            # Check cancel
            if self._cancel_event.is_set():
                self.em.log("WARN", "Cancel requested – stopping cell-by-cell.")
                break
            # Check pause (blocks without busy loop)
            self._pause_event.wait()

            ws = wb[sheet_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            original = str(cell.value)
            try:
                translated = self._translate_cell(original)
                self._write_cell_result(cell, original, translated)
                self._translated_count += 1
            except Exception as exc:
                tb = traceback.format_exc()
                self.em.log(
                    "ERROR",
                    f"Cell {cell_id} failed: {type(exc).__name__}: {exc}",
                )
                self.em.log("ERROR", f"Traceback:\n{tb}")
                self._error_count += 1
                self._error_log.append((sheet_name, cell_id, original, "", f"{type(exc).__name__}: {exc}"))
            done += 1
            self.em.progress(done, total)
            if self.min_interval > 0 and done < total:
                time.sleep(self.min_interval)

    def _translate_cell(self, text: str) -> str:
        if self.use_glossary and self._term_override:
            text = self._term_override.pre_replace(text)
        mapping: Dict[str, str] = {}
        if self.use_protection:
            text, mapping = regex_protection.protect(text)
        translated = self.translator.translate(
            text, self.source_lang, self.target_lang
        )
        if self.use_protection and mapping:
            translated = regex_protection.restore(translated, mapping)
        if self.use_glossary and self._term_override:
            translated = self._term_override.post_fix(translated)
        return translated

    # ------------------------------------------------------------------
    # Output mode cell writer
    # ------------------------------------------------------------------

    @staticmethod
    def _hex_to_argb(color_hex: str) -> str:
        text = (color_hex or "").strip().upper()
        if not text:
            return "FF000000"
        if text.startswith("#"):
            text = text[1:]
        if re.fullmatch(r"[0-9A-F]{6}", text):
            return f"FF{text}"
        return "FF000000"

    @staticmethod
    def _font_color_to_argb(color) -> str:
        """Convert openpyxl Color object to ARGB hex string."""
        if color is None:
            return "FF000000"
        if hasattr(color, "rgb") and color.rgb:
            c = str(color.rgb).strip().upper().lstrip("#")
            if len(c) == 6:
                return f"FF{c}"
            if len(c) == 8 and c != "00000000":
                return c
        return "FF000000"

    @staticmethod
    def _preload_cell_fonts(xlsx_path) -> dict:
        """Read original per-cell font info from raw XLSX XML.

        Returns {(sheet_name, coord): {name, sz, color, b, i}}.
        Reads BOTH:
          1. Style-based fonts (styles.xml + theme1.xml)
          2. Rich-text run fonts (<rPr> in inline strings / shared strings)
        """
        import zipfile
        import xml.etree.ElementTree as ET
        NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
        result: dict = {}
        try:
            with zipfile.ZipFile(str(xlsx_path), "r") as z:
                names = z.namelist()

                # --- 1. Read theme colors ---
                theme_colors: list[str] = []
                if "xl/theme/theme1.xml" in names:
                    theme_root = ET.fromstring(z.read("xl/theme/theme1.xml"))
                    clrScheme = theme_root.find(f".//{{{A_NS}}}clrScheme")
                    if clrScheme is not None:
                        for child in clrScheme:
                            for sub in child:
                                val = sub.get("val", sub.get("lastClr", "000000"))
                                if val.lower() == "windowtext":
                                    val = "000000"
                                elif val.lower() == "window":
                                    val = "FFFFFF"
                                theme_colors.append(val.upper())
                    # OOXML theme indices are swapped: 0↔1, 2↔3
                    # XML order: dk1,lt1,dk2,lt2,accent1...
                    # Theme idx:  1,  0,  3,  2,    4...
                    if len(theme_colors) >= 4:
                        theme_colors[0], theme_colors[1] = theme_colors[1], theme_colors[0]
                        theme_colors[2], theme_colors[3] = theme_colors[3], theme_colors[2]

                def _resolve_color(color_el) -> str:
                    """Convert <color> element to FFRRGGBB."""
                    if color_el is None:
                        return "FF000000"
                    rgb = color_el.get("rgb")
                    if rgb and rgb != "00000000":
                        return rgb if len(rgb) == 8 else f"FF{rgb}"
                    theme = color_el.get("theme")
                    if theme is not None:
                        ti = int(theme)
                        if ti < len(theme_colors):
                            return f"FF{theme_colors[ti]}"
                    indexed = color_el.get("indexed")
                    if indexed is not None:
                        return _indexed_to_argb(indexed)
                    return "FF000000"

                # --- 2. Read styles.xml (fonts + cellXfs) ---
                style_fonts: list[dict] = []
                xf_to_font: list[int] = []
                if "xl/styles.xml" in names:
                    styles_root = ET.fromstring(z.read("xl/styles.xml"))
                    fonts_el = styles_root.find(f"{{{NS}}}fonts")
                    if fonts_el is not None:
                        for font_el in fonts_el:
                            name_el = font_el.find(f"{{{NS}}}name")
                            sz_el = font_el.find(f"{{{NS}}}sz")
                            color_el = font_el.find(f"{{{NS}}}color")
                            b_el = font_el.find(f"{{{NS}}}b")
                            i_el = font_el.find(f"{{{NS}}}i")
                            style_fonts.append({
                                "name": name_el.get("val", "Calibri") if name_el is not None else "Calibri",
                                "sz": float(sz_el.get("val", "11")) if sz_el is not None else 11,
                                "color": _resolve_color(color_el),
                                "b": b_el is not None,
                                "i": i_el is not None,
                            })
                    xfs_el = styles_root.find(f"{{{NS}}}cellXfs")
                    if xfs_el is not None:
                        for xf in xfs_el:
                            xf_to_font.append(int(xf.get("fontId", "0")))

                # --- 3. Shared strings: store ALL runs per entry ---
                ss_runs: dict[int, list] = {}  # si_idx → [(text, font_dict), ...]
                if "xl/sharedStrings.xml" in names:
                    ss_root = ET.fromstring(z.read("xl/sharedStrings.xml"))
                    for idx, si in enumerate(list(ss_root)):
                        r_els = si.findall(f"{{{NS}}}r")
                        if r_els:
                            runs_list = []
                            for r_el in r_els:
                                t_el = r_el.find(f"{{{NS}}}t")
                                text = t_el.text if t_el is not None else ""
                                rPr = r_el.find(f"{{{NS}}}rPr")
                                font = _parse_rPr(rPr, NS, theme_colors) if rPr is not None else {"name": "Calibri", "sz": 11, "color": "FF000000", "b": False, "i": False}
                                runs_list.append((text, font))
                            ss_runs[idx] = runs_list

                # --- 4. Map sheet name → file ---
                wb_xml = z.read("xl/workbook.xml") if "xl/workbook.xml" in names else None
                if not wb_xml:
                    return result
                wb_root = ET.fromstring(wb_xml)
                rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
                rels_xml = z.read("xl/_rels/workbook.xml.rels") if "xl/_rels/workbook.xml.rels" in names else b""
                rid_to_target = {}
                if rels_xml:
                    for rel in ET.fromstring(rels_xml):
                        rid_to_target[rel.get("Id", "")] = rel.get("Target", "")
                sname_to_file: dict[str, str] = {}
                for s_el in wb_root.findall(f".//{{{NS}}}sheet"):
                    sname = s_el.get("name", "")
                    rid = s_el.get(f"{{{rel_ns}}}id", "")
                    tgt = rid_to_target.get(rid, "")
                    if tgt:
                        sname_to_file[sname] = f"xl/{tgt}" if not tgt.startswith("/") else tgt.lstrip("/")

                # --- 5. Read each sheet's cells ---
                # Value is a list of (text, font_dict) tuples per cell.
                for sname, sfile in sname_to_file.items():
                    if sfile not in names:
                        continue
                    sheet_root = ET.fromstring(z.read(sfile))
                    for c_el in sheet_root.iter(f"{{{NS}}}c"):
                        coord = c_el.get("r", "")
                        ctype = c_el.get("t", "")
                        s_idx = int(c_el.get("s", "0"))
                        runs_data = None

                        # Priority 1: rich text runs (ALL runs, not just first)
                        if ctype == "inlineStr":
                            is_el = c_el.find(f"{{{NS}}}is")
                            if is_el is not None:
                                r_els = is_el.findall(f"{{{NS}}}r")
                                if r_els:
                                    runs_data = []
                                    for r_el in r_els:
                                        t_el = r_el.find(f"{{{NS}}}t")
                                        text = t_el.text if t_el is not None else ""
                                        rPr = r_el.find(f"{{{NS}}}rPr")
                                        font = _parse_rPr(rPr, NS, theme_colors) if rPr is not None else {"name": "Calibri", "sz": 11, "color": "FF000000", "b": False, "i": False}
                                        runs_data.append((text, font))
                        elif ctype == "s":
                            v_el = c_el.find(f"{{{NS}}}v")
                            if v_el is not None and v_el.text is not None:
                                si_idx = int(v_el.text)
                                runs_data = ss_runs.get(si_idx)

                        # Priority 2: cell style font (single run)
                        if runs_data is None and style_fonts and xf_to_font:
                            if s_idx < len(xf_to_font):
                                font_id = xf_to_font[s_idx]
                                if font_id < len(style_fonts):
                                    runs_data = [(None, dict(style_fonts[font_id]))]  # text=None means use cell.value

                        if runs_data:
                            result[(sname, coord)] = runs_data
        except Exception:
            pass  # fallback to cell.font
        return result

    def _apply_custom_cell_font(self, cell) -> None:
        from openpyxl.styles import Font

        old_font = cell.font or Font()
        cell.font = Font(
            name=self._custom_font_name or old_font.name,
            size=self._custom_font_size,
            bold=self._custom_bold,
            italic=self._custom_italic,
            underline="single" if self._custom_underline else None,
            color=self._hex_to_argb(self._custom_font_color),
            charset=old_font.charset,
            family=old_font.family,
            scheme=old_font.scheme,
            vertAlign=old_font.vertAlign,
            strike=old_font.strike,
            outline=old_font.outline,
            shadow=old_font.shadow,
            condense=old_font.condense,
            extend=old_font.extend,
        )

    @staticmethod
    def _font_color_to_argb(color_obj) -> str | None:
        if color_obj is None:
            return None
        try:
            rgb = getattr(color_obj, "rgb", None)
            if rgb:
                text = str(rgb).strip().upper().lstrip("#")
                if re.fullmatch(r"[0-9A-F]{6}", text):
                    return f"FF{text}"
                if re.fullmatch(r"[0-9A-F]{8}", text) and text != "00000000":
                    return text
            # Handle indexed colors (e.g. indexed=8 → black, indexed=10 → red)
            indexed = getattr(color_obj, "indexed", None)
            if indexed is not None:
                return _indexed_to_argb(str(indexed))
            # Theme colors are already resolved by preloaded fonts in most cases;
            # this is a last-resort fallback.
            theme = getattr(color_obj, "theme", None)
            if theme is not None:
                # Without access to theme_colors list here, fall back to black
                # for theme=1 (dark text) and white for theme=0.
                ti = int(theme)
                if ti == 1:
                    return "FF000000"
                elif ti == 0:
                    return "FFFFFFFF"
        except Exception:
            return None
        return None

    def _to_inline_font(self, font):
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles import Font

        f = font or Font()
        return InlineFont(
            rFont=f.name,
            charset=f.charset,
            family=f.family,
            b=f.bold,
            i=f.italic,
            strike=f.strike,
            outline=f.outline,
            shadow=f.shadow,
            condense=f.condense,
            extend=f.extend,
            color=self._font_color_to_argb(f.color),
            sz=f.size,
            u=f.underline,
            vertAlign=f.vertAlign,
            scheme=f.scheme,
        )

    def _custom_inline_font(self, base_font):
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles import Font

        base = base_font or Font()
        return InlineFont(
            rFont=self._custom_font_name or base.name,
            charset=base.charset,
            family=base.family,
            b=self._custom_bold,
            i=self._custom_italic,
            strike=base.strike,
            outline=base.outline,
            shadow=base.shadow,
            condense=base.condense,
            extend=base.extend,
            color=self._hex_to_argb(self._custom_font_color),
            sz=self._custom_font_size,
            u="single" if self._custom_underline else None,
            vertAlign=base.vertAlign,
            scheme=base.scheme,
        )

    @staticmethod
    def _segments_differ(seg1, seg2) -> bool:
        """Return True if two segment lists have different text content."""
        if not seg1 or not seg2:
            return False
        text1 = "".join(getattr(b, 'text', str(b)) for b in seg1)
        text2 = "".join(getattr(b, 'text', str(b)) for b in seg2)
        return text1.strip() != text2.strip()

    def _write_split_rich_text(self, cell, original: str, translated: str) -> bool:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont as _InlineFont

        source_font = self._to_inline_font(cell.font)
        if self._text_format_mode == "custom_format":
            target_font = self._custom_inline_font(cell.font)
        else:
            target_font = self._to_inline_font(cell.font)

        source_segment = self._segment_rich_text(original, self.source_lang, source_font, cell.font)
        target_segment = self._segment_rich_text(translated, self.target_lang, target_font, cell.font)

        # Use a neutral font for separator newlines between blocks
        sep_font = source_font

        blocks: list[TextBlock] = []

        if self.output_mode == "prefix":
            if target_segment:
                blocks.extend(target_segment)
            # Only append original if different from translation (avoid duplicate for untranslated text)
            if source_segment and self._segments_differ(target_segment, source_segment):
                blocks.append(TextBlock(sep_font, "\n"))
                blocks.extend(source_segment)
        elif self.output_mode == "suffix":
            if source_segment:
                blocks.extend(source_segment)
            if target_segment and self._segments_differ(source_segment, target_segment):
                blocks.append(TextBlock(sep_font, "\n"))
                blocks.extend(target_segment)
        elif self.output_mode in ("overwrite", "new_sheet"):
            if target_segment:
                blocks.extend(target_segment)
        else:
            if target_segment:
                blocks.extend(target_segment)

        if blocks:
            cell.value = CellRichText(blocks)
            # Reset cell-level font to default so InlineFont in each TextBlock
            # takes full effect.
            from openpyxl.styles import Font as _Font
            cell.font = _Font()
        return True

    def _write_cell_result(self, cell, original: str, translated: str) -> None:
        """Write translation into cell.

        Source text keeps its original font.
        Translated text uses the user-chosen custom font/size/color.
        Pinyin uses the user-chosen pinyin font/size/color.

        Plain text is written first (openpyxl saves as shared strings).
        Font metadata is stored in _rich_text_records for XLSX post-processing.
        """
        from openpyxl.styles import Alignment

        original = (original or "").replace("\r", "")
        translated = (translated or "").replace("\r", "")

        # Deduplicate identical lines in prefix/suffix mode to avoid duplicating untranslated lines
        if self.output_mode in ("prefix", "suffix") and '\n' in original:
            orig_lines = [l.strip().lower() for l in original.split('\n')]
            trans_lines = translated.split('\n')
            new_trans_lines = []
            for line in trans_lines:
                if line.strip().lower() in orig_lines:
                    continue
                new_trans_lines.append(line)
            if new_trans_lines:
                translated = '\n'.join(new_trans_lines)
            else:
                translated = original

        # --- collect font info from source cell --------------------------
        # Use pre-loaded original runs from raw XLSX XML (accurate per-run fonts)
        cf = cell.font
        sheet_name = cell.parent.title
        orig_fonts = getattr(self, '_original_cell_fonts', {})
        preloaded_runs = None  # list of (text, font_dict)

        # Try current sheet name first, then strip lang suffixes recursively
        preloaded_runs = orig_fonts.get((sheet_name, cell.coordinate))
        if preloaded_runs is None:
            _SUFFIXES = ['_Vi', '_En', '_Ja', '_Ko', '_Fr', '_De', '_Es', '_Pt', '_Ru',
                         '_Zh', '_Th', '_Id', '_Ms', '_Ar', '_Hi', '_Tr']
            name = sheet_name
            while preloaded_runs is None:
                stripped = False
                for suffix in _SUFFIXES:
                    if name.endswith(suffix):
                        name = name[:-len(suffix)]
                        preloaded_runs = orig_fonts.get((name, cell.coordinate))
                        stripped = True
                        break
                if not stripped:
                    break

        # Build source runs: use preloaded per-run fonts if available
        fallback_font = {
            "name": cf.name or "Calibri",
            "sz": cf.size or 11,
            "color": self._font_color_to_argb(cf.color) or "FF000000",
            "b": bool(cf.bold),
            "i": bool(cf.italic),
        }

        if preloaded_runs:
            # Check if this is a multi-run cell (rich text from previous translation)
            if len(preloaded_runs) > 1:
                # Use the original per-run data directly as src_runs
                src_runs = [(t or "", dict(f)) for t, f in preloaded_runs]
                # Also need src_font_info for newline/separator
                src_font_info = dict(preloaded_runs[0][1])
            else:
                # Single run or style-based font
                src_font_info = dict(preloaded_runs[0][1])
                src_runs = None  # will be built below
        else:
            src_font_info = fallback_font
            src_runs = None

        tgt_font_info = dict(src_font_info)
        if self._text_format_mode == "custom_format":
            if self._custom_font_name != "Mặc định":
                tgt_font_info["name"] = self._custom_font_name
            if self._custom_font_size > 0:
                tgt_font_info["sz"] = self._custom_font_size
            if self._custom_font_color != "default":
                tgt_font_info["color"] = self._hex_to_argb(self._custom_font_color)
            if self._custom_bold != "default":
                tgt_font_info["b"] = bool(self._custom_bold)
            if self._custom_italic != "default":
                tgt_font_info["i"] = bool(self._custom_italic)
            if hasattr(self, '_custom_underline') and self._custom_underline != "default":
                tgt_font_info["u"] = bool(self._custom_underline)
        py_font_info = {
            "name": self._pinyin_font_family or "Arial",
            "sz": self._pinyin_font_size or 9,
            "color": self._hex_to_argb(self._pinyin_font_color) or "FF888888",
            "b": False,
            "i": True,
        } if self._add_chinese_pinyin else dict(src_font_info)

        # --- build runs: list of (text, font_info) -----------------------
        def _build_runs(text, lang, font):
            runs = []
            for line in text.split("\n"):
                if runs:
                    runs.append(("\n", font))
                if line:
                    runs.append((line, font))
                py = self._pinyin_for_lang(line, lang)
                if py:
                    runs.append(("\n", py_font_info))
                    runs.append((py, py_font_info))
            return runs

        if src_runs is None:
            src_runs = _build_runs(original, self.source_lang, src_font_info)
        tgt_runs = _build_runs(translated, self.target_lang, tgt_font_info)

        all_runs = []
        if self.output_mode == "prefix":
            all_runs.extend(tgt_runs)
            if src_runs and translated != original:
                all_runs.append(("\n", src_font_info))
                all_runs.extend(src_runs)
        elif self.output_mode == "suffix":
            all_runs.extend(src_runs)
            if tgt_runs and translated != original:
                all_runs.append(("\n", tgt_font_info))
                all_runs.extend(tgt_runs)
        else:
            all_runs.extend(tgt_runs)

        # --- write plain text to cell ------------------------------------
        plain_text = "".join(t for t, _ in all_runs)
        cell.value = plain_text

        # --- store metadata for post-processing --------------------------
        if not hasattr(self, "_rich_text_records"):
            self._rich_text_records = []
        self._rich_text_records.append({
            "sheet": cell.parent.title,
            "coord": cell.coordinate,
            "plain": plain_text,
            "runs": [(t, dict(f)) for t, f in all_runs],
        })

        # --- enable wrap-text so \n renders as line breaks ---------------
        old_align = cell.alignment or Alignment()
        cell.alignment = Alignment(
            horizontal=old_align.horizontal,
            vertical=old_align.vertical,
            wrap_text=True,
            shrink_to_fit=old_align.shrink_to_fit,
            indent=old_align.indent,
            text_rotation=old_align.text_rotation,
        )


    # ------------------------------------------------------------------
    # Post-processing: convert plain text → rich text in sharedStrings.xml
    # ------------------------------------------------------------------

    def _post_process_rich_text(self, xlsx_path: Path) -> None:
        """Rewrite XLSX so translated cells have rich-text font runs.

        openpyxl saves text on copied worksheets as inline strings
        (`<is><t>text</t></is>`) which do NOT support per-run styling
        in Excel.  This method converts them into shared-string
        references (`<si><r>...</r></si>`) which Excel renders correctly.
        """
        import zipfile
        import xml.etree.ElementTree as ET

        NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

        # Register all OOXML namespaces to prevent ns0/ns1 prefix corruption
        ET.register_namespace("", NS)
        ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")
        ET.register_namespace("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006")
        ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")
        ET.register_namespace("xr", "http://schemas.microsoft.com/office/spreadsheetml/2014/revision")
        ET.register_namespace("xr6", "http://schemas.microsoft.com/office/spreadsheetml/2014/revision6")
        ET.register_namespace("xr10", "http://schemas.microsoft.com/office/spreadsheetml/2014/revision10")
        ET.register_namespace("xml", "http://www.w3.org/XML/1998/namespace")
        records = getattr(self, "_rich_text_records", [])
        if not records:
            return

        # --- Read all zip contents ---
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            names = zin.namelist()
            contents = {n: zin.read(n) for n in names}

        # --- Map sheet name → worksheet file ---
        wb_xml = contents.get("xl/workbook.xml")
        if not wb_xml:
            return
        wb_root = ET.fromstring(wb_xml)
        rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        rels_xml = contents.get("xl/_rels/workbook.xml.rels", b"")
        rid_to_target = {}
        if rels_xml:
            for rel in ET.fromstring(rels_xml):
                rid_to_target[rel.get("Id", "")] = rel.get("Target", "")
        sheet_name_to_file: dict[str, str] = {}
        for s_el in wb_root.findall(f".//{{{NS}}}sheet"):
            sname = s_el.get("name", "")
            rid = s_el.get(f"{{{rel_ns}}}id", "")
            tgt = rid_to_target.get(rid, "")
            if tgt:
                sheet_name_to_file[sname] = f"xl/{tgt}" if not tgt.startswith("/") else tgt.lstrip("/")

        # --- Build record lookup: (sheet_file, coord) → runs ---
        coord_to_runs: dict[tuple[str, str], list] = {}
        for rec in records:
            sf = sheet_name_to_file.get(rec["sheet"])
            if sf:
                coord_to_runs[(sf, rec["coord"])] = rec["runs"]

        # --- Also restore rich text on EXISTING sheets that openpyxl corrupted ---
        # When openpyxl saves, it converts shared strings → inline strings,
        # destroying rich text on sheets like _Vi. Restore them using preloaded data.
        orig_fonts = getattr(self, "_original_cell_fonts", {})
        for (sname, coord), runs_data in orig_fonts.items():
            # Skip cells with text=None (style-based, no rich text)
            if not isinstance(runs_data, list):
                continue
            if any(t is None for t, _ in runs_data):
                continue
            sf = sheet_name_to_file.get(sname)
            if sf and (sf, coord) not in coord_to_runs:
                coord_to_runs[(sf, coord)] = [(t or "", dict(f)) for t, f in runs_data]

        if not coord_to_runs:
            return

        # --- Parse or create sharedStrings.xml ---
        if "xl/sharedStrings.xml" in contents:
            ss_root = ET.fromstring(contents["xl/sharedStrings.xml"])
        else:
            ss_root = ET.Element("sst")
            ss_root.set("xmlns", NS)

        # Count existing shared strings
        existing_si = list(ss_root)
        next_idx = len(existing_si)

        # --- Helper: build a rich-text <si> element ---
        def _make_rich_si(runs):
            si = ET.Element("si")
            for text, font in runs:
                r_el = ET.SubElement(si, "r")
                rPr = ET.SubElement(r_el, "rPr")
                if font.get("b"):
                    ET.SubElement(rPr, "b")
                if font.get("i"):
                    ET.SubElement(rPr, "i")
                sz_el = ET.SubElement(rPr, "sz")
                sz_el.set("val", str(font.get("sz", 11)))
                color_el = ET.SubElement(rPr, "color")
                color_el.set("rgb", str(font.get("color", "FF000000")))
                rFont_el = ET.SubElement(rPr, "rFont")
                rFont_el.set("val", str(font.get("name", "Calibri")))
                t_el = ET.SubElement(r_el, "t")
                t_el.text = text
                t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            return si

        # --- Process each worksheet ---
        modified_sheets: set[str] = set()
        for sheet_file in {sf for sf, _ in coord_to_runs}:
            if sheet_file not in contents:
                continue
            sheet_root = ET.fromstring(contents[sheet_file])

            for c_el in sheet_root.iter(f"{{{NS}}}c"):
                ref = c_el.get("r", "")
                key = (sheet_file, ref)
                if key not in coord_to_runs:
                    continue

                runs = coord_to_runs[key]

                # Create rich-text shared string entry
                rich_si = _make_rich_si(runs)
                ss_root.append(rich_si)
                si_idx = next_idx
                next_idx += 1

                # Convert cell to shared string reference
                c_el.set("t", "s")

                # Remove old <is> or <v>
                for child_tag in ["is", "v"]:
                    old = c_el.find(f"{{{NS}}}{child_tag}")
                    if old is not None:
                        c_el.remove(old)

                # Add new <v> pointing to the shared string
                v_el = ET.SubElement(c_el, "v")
                v_el.text = str(si_idx)

                modified_sheets.add(sheet_file)

            if sheet_file in modified_sheets:
                contents[sheet_file] = ET.tostring(
                    sheet_root, xml_declaration=True, encoding="UTF-8"
                )

        if not modified_sheets:
            return

        # --- Update shared string counts ---
        ss_root.set("count", str(next_idx))
        ss_root.set("uniqueCount", str(next_idx))
        contents["xl/sharedStrings.xml"] = ET.tostring(
            ss_root, xml_declaration=True, encoding="UTF-8"
        )

        # --- Ensure [Content_Types].xml includes sharedStrings ---
        ct_key = "[Content_Types].xml"
        if ct_key in contents:
            ct_bytes = contents[ct_key]
            if b"sharedStrings" not in ct_bytes:
                # Inject Override element before closing </Types>
                snippet = b'<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                ct_bytes = ct_bytes.replace(b"</Types>", snippet + b"</Types>")
                contents[ct_key] = ct_bytes

        # --- Ensure workbook.xml.rels references sharedStrings ---
        rels_key = "xl/_rels/workbook.xml.rels"
        if rels_key in contents:
            rels_bytes = contents[rels_key]
            if b"sharedStrings" not in rels_bytes:
                # Find max rId by simple regex
                import re
                ids = [int(m) for m in re.findall(rb'rId(\d+)', rels_bytes)]
                max_id = max(ids) if ids else 0
                snippet = (
                    f'<Relationship Id="rId{max_id+1}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
                    f'Target="sharedStrings.xml"/>'
                ).encode()
                rels_bytes = rels_bytes.replace(b"</Relationships>", snippet + b"</Relationships>")
                contents[rels_key] = rels_bytes

        # --- Rewrite the XLSX zip ---
        with zipfile.ZipFile(str(xlsx_path), "w", zipfile.ZIP_DEFLATED) as zout:
            for name in names:
                zout.writestr(name, contents[name])
            if "xl/sharedStrings.xml" not in names:
                zout.writestr("xl/sharedStrings.xml", contents["xl/sharedStrings.xml"])

    # ------------------------------------------------------------------
    # Safe save
    # ------------------------------------------------------------------

    def _safe_save(self, wb: Workbook, out_path: Path) -> Path:
        """Save workbook, auto-renaming if file is locked (PermissionError)."""
        for attempt in range(10):
            target = out_path if attempt == 0 else (
                out_path.parent / f"{out_path.stem}({attempt}){out_path.suffix}"
            )
            try:
                wb.save(str(target))
                if attempt > 0:
                    self.em.log(
                        "WARN",
                        f"File '{out_path.name}' was locked, saved as '{target.name}'",
                    )
                wb.close()
                return target
            except PermissionError:
                self.em.log(
                    "WARN",
                    f"PermissionError on '{target.name}', trying next name…",
                )
        # All attempts failed
        self.em.log("ERROR", f"Cannot save after 10 attempts: {out_path.name}")
        wb.close()
        return out_path

    # ── OCR Image helpers ─────────────────────────────────────────

    def _collect_xlsx_images(self, wb, sheet_names: list) -> list:
        """Collect all images from specified sheets for OCR."""
        images = []
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for img in ws._images:
                try:
                    image_bytes = img._data()
                    # Store: (image_bytes, sheet_name, anchor_col, anchor_row)
                    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
                    col = 0
                    row = 0
                    if hasattr(img, 'anchor') and img.anchor:
                        if hasattr(img.anchor, '_from'):
                            col = img.anchor._from.col
                            row = img.anchor._from.row
                        elif hasattr(img.anchor, 'col'):
                            col = img.anchor.col
                            row = img.anchor.row
                    images.append((image_bytes, sheet_name, col, row))
                except Exception:
                    pass
        return images

    def _process_xlsx_images(self, images: list, wb) -> None:
        """OCR, translate, and write results to _OCR_Results sheet."""
        self.em.log("INFO", f"Starting OCR for {len(images)} images in Excel workbook")

        # Build ocr input
        ocr_input = [(None, None, img[0], i) for i, img in enumerate(images)]

        # Step 1: OCR
        ocr_results = self.ocr_pipeline.ocr_and_filter_images(ocr_input, self.source_lang)

        # Step 2: Translate
        self.ocr_pipeline.batch_translate_segments(ocr_results, self.source_lang, self.target_lang)

        # Step 3: Create results sheet
        sheet_name = "_OCR_Results"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_ocr = wb.create_sheet(sheet_name)
        ws_ocr.append(["Image #", "Block #", "Sheet", "Cell", "Original Text", "Translated Text"])

        for idx, result_item in enumerate(ocr_results):
            result = result_item.get("result")
            if not result or result.error or not result.blocks:
                continue

            _, sheet, col, row = images[idx]
            cell_ref = f"{chr(65+col) if col < 26 else '?'}{row+1}"

            block_num = 0
            for block in result.blocks:
                if not block.text or not block.text.strip():
                    continue
                if getattr(block, "skipped", False):
                    continue
                original = block.text.strip()
                translated = (block.translated_text or "").strip()
                if not translated:
                    translated = original
                ws_ocr.append([idx+1, block_num+1, sheet, cell_ref, original, translated])
                block_num += 1

        self.em.log("INFO", f"OCR results written to '{sheet_name}' sheet")

    def _duplicate_sheets_via_com(self, file_path: Path, sheets_todo: list[str], target_lang: str) -> tuple[Path, dict[str, str]]:
        """
        Duplicate target sheets using COM to preserve all formatting, charts, and media.
        """
        import win32com.client
        import pythoncom
        
        temp_path = file_path.parent / f"~temp_dup_{file_path.name}"
        if temp_path.exists():
            try:
                temp_path.unlink()
            except Exception:
                pass
                
        pythoncom.CoInitialize()
        excel = None
        sheet_mapping = {}
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            wb = excel.Workbooks.Open(str(file_path.resolve()))
            
            existing_names = [wb.Sheets(i).Name for i in range(1, wb.Sheets.Count + 1)]
            
            for orig_name in sheets_todo:
                trans_name = f"{orig_name}_{target_lang[:2].upper()}"
                if len(trans_name) > 31:
                    trans_name = trans_name[:31]
                
                counter = 2
                base_trans = trans_name
                while trans_name in existing_names:
                    suffix = f"_{counter}"
                    trans_name = base_trans[:31 - len(suffix)] + suffix
                    counter += 1
                
                existing_names.append(trans_name)
                sheet_mapping[orig_name] = trans_name
                
                src_sheet = wb.Sheets(orig_name)
                # Copy sheet and place it after the original
                src_sheet.Copy(After=src_sheet)
                copied_sheet = wb.Sheets(src_sheet.Index + 1)
                copied_sheet.Name = trans_name
                
                # Change tab color to blue
                try:
                    copied_sheet.Tab.Color = 0xFF0000  # Blue
                except Exception:
                    pass
            
            wb.SaveAs(str(temp_path.resolve()))
            wb.Close(SaveChanges=True)
        except Exception as exc:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass
            raise exc
        finally:
            if excel:
                excel.Quit()
            pythoncom.CoUninitialize()
            
        return temp_path, sheet_mapping

    def _copy_images(self, source_ws, target_ws) -> None:
        """Copy images from source to target worksheet (openpyxl does not do this)."""
        try:
            for img in source_ws._images:
                new_img = deepcopy(img)
                target_ws.add_image(new_img)
        except Exception as exc:
            self.em.log("WARN", f"Image copy failed: {type(exc).__name__}: {exc}")
    # ------------------------------------------------------------------
    # Cell scanning
    # ------------------------------------------------------------------

    def _scan_cells(
        self, wb: Workbook, sheet_names: Optional[list[str]] = None,
    ) -> dict:
        """Scan specified sheets for translatable cells."""
        cells_map = {}
        names = sheet_names or wb.sheetnames
        for sheet_name in names:
            ws = wb[sheet_name]
            # Determine start row for this sheet
            # sheet_target_rows uses ORIGINAL sheet names, but translated sheet
            # names may have a suffix like _Vi. Try to look up both.
            min_row = 1
            if self.sheet_target_rows:
                # Direct lookup
                if sheet_name in self.sheet_target_rows:
                    min_row = self.sheet_target_rows[sheet_name]
                else:
                    # Try to find matching original name (strip suffix)
                    for orig_name, start_r in self.sheet_target_rows.items():
                        if sheet_name.startswith(orig_name):
                            min_row = start_r
                            break
            if min_row > 1:
                self.em.log("INFO", f"[scan] {sheet_name}: starting from row {min_row}")
            for row_idx, row in enumerate(ws.iter_rows(min_row=min_row), min_row):
                if row_idx > self.max_rows:
                    break
                for col_idx, cell in enumerate(row, 1):
                    if col_idx > self.max_cols:
                        break
                    if self._is_eligible(cell):
                        cell_id = f"{sheet_name}!{cell.coordinate}"
                        cells_map[cell_id] = (sheet_name, row_idx, col_idx)
        return cells_map

    def _count_all_cells(self, wb: Workbook, sheet_names: list[str]) -> int:
        """Count all non-empty cells in the specified sheets."""
        count = 0
        for name in sheet_names:
            ws = wb[name]
            for row in ws.iter_rows(min_row=1, max_row=self.max_rows):
                for cell in row:
                    if cell.value is not None:
                        count += 1
        return count

    def _is_eligible(self, cell) -> bool:
        val = cell.value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return False
        if isinstance(val, (datetime, dt_date, dt_time)):
            return False
        if getattr(cell, "is_date", False):
            return False
        if isinstance(val, str) and self._looks_like_datetime_text(val):
            return False
        # Skip formulas
        if isinstance(val, str) and val.startswith("="):
            return False
        if self.skip_numeric:
            if isinstance(val, (int, float)):
                return False
            if isinstance(val, str) and re.fullmatch(r"[\d.,\s]+", val):
                return False
        return True

    @staticmethod
    def _looks_like_datetime_text(text: str) -> bool:
        candidate = (text or "").strip()
        if not candidate:
            return False
        return any(pattern.fullmatch(candidate) for pattern in _DATETIME_TEXT_PATTERNS)
