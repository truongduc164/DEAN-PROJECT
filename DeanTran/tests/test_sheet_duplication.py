"""
tests/test_sheet_duplication.py – Tests for sheet duplication, tab color, and naming.

Covers:
  1. Translated sheet exists with correct name
  2. Tab color is blue
  3. Original sheet values unchanged
  4. Long sheet name truncation (31-char limit)
  5. Multiple sheets handled correctly
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from app.core.translators.translator_service import MockTranslator
from app.core.excel_processor import (
    ExcelProcessor, build_output_path, _make_translated_sheet_name, BLUE_TAB_COLOR,
)
from app.core.event_manager import EventManager


def _create_multi_sheet_workbook(path: Path) -> None:
    """Create a workbook with multiple sheets."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Project overview"
    ws1["A2"] = "Status report"

    ws2 = wb.create_sheet("Details")
    ws2["A1"] = "Item one"
    ws2["B1"] = "Description here"
    ws2["A2"] = 100  # numeric, skip

    wb.save(str(path))


def _create_long_name_workbook(path: Path) -> None:
    """Create a workbook with a very long sheet name."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ThisIsAVeryLongSheetNameThatIs"  # 30 chars
    ws["A1"] = "Test data"
    wb.save(str(path))


class TestSheetDuplication:
    def test_translated_sheet_exists(self, tmp_path: Path):
        src = tmp_path / "multi.xlsx"
        _create_multi_sheet_workbook(src)

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="en",
            target_lang="Vietnamese",
        )
        out = proc.process(src)
        wb = openpyxl.load_workbook(str(out))

        assert "Summary" in wb.sheetnames
        assert "Summary_Vi" in wb.sheetnames
        assert "Details" in wb.sheetnames
        assert "Details_Vi" in wb.sheetnames

    def test_tab_color_is_blue(self, tmp_path: Path):
        src = tmp_path / "color.xlsx"
        _create_multi_sheet_workbook(src)

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="en",
            target_lang="Vietnamese",
        )
        out = proc.process(src)
        wb = openpyxl.load_workbook(str(out))

        # Translated sheets should have blue tab color
        ws_vi = wb["Summary_Vi"]
        tab_color = ws_vi.sheet_properties.tabColor
        assert tab_color is not None
        # openpyxl stores color as Color object; check rgb value
        assert tab_color.rgb is not None
        assert BLUE_TAB_COLOR.lower() in tab_color.rgb.lower()

    def test_original_sheet_unchanged(self, tmp_path: Path):
        src = tmp_path / "unchanged.xlsx"
        _create_multi_sheet_workbook(src)

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="en",
            target_lang="Vietnamese",
        )
        out = proc.process(src)
        wb = openpyxl.load_workbook(str(out))

        # Original "Summary" sheet values must be preserved
        ws_orig = wb["Summary"]
        assert ws_orig["A1"].value == "Project overview"
        assert ws_orig["A2"].value == "Status report"

        # Original "Details" sheet
        ws_det = wb["Details"]
        assert ws_det["A1"].value == "Item one"
        assert ws_det["A2"].value == 100

    def test_translated_sheet_has_translations(self, tmp_path: Path):
        src = tmp_path / "trans.xlsx"
        _create_multi_sheet_workbook(src)

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="en",
            target_lang="Vietnamese",
        )
        out = proc.process(src)
        wb = openpyxl.load_workbook(str(out))

        ws_vi = wb["Summary_Vi"]
        assert ws_vi["A1"].value == "[Vietnamese] Project overview"
        assert ws_vi["A2"].value == "[Vietnamese] Status report"

        ws_det = wb["Details_Vi"]
        assert ws_det["A1"].value == "[Vietnamese] Item one"
        assert ws_det["B1"].value == "[Vietnamese] Description here"
        # Numeric should be unchanged
        assert ws_det["A2"].value == 100

    def test_english_target(self, tmp_path: Path):
        src = tmp_path / "eng.xlsx"
        _create_multi_sheet_workbook(src)

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="zh",
            target_lang="English",
        )
        out = proc.process(src)
        assert out.name == "eng_En.xlsx"

        wb = openpyxl.load_workbook(str(out))
        assert "Summary_En" in wb.sheetnames
        assert "Details_En" in wb.sheetnames

        ws = wb["Summary_En"]
        assert ws["A1"].value == "[English] Project overview"


class TestSheetNameTruncation:
    def test_make_translated_sheet_name_normal(self):
        name = _make_translated_sheet_name("Sheet1", "Vietnamese", set())
        assert name == "Sheet1_Vi"
        assert len(name) <= 31

    def test_make_translated_sheet_name_long(self):
        long_name = "A" * 30  # 30 chars
        name = _make_translated_sheet_name(long_name, "Vietnamese", set())
        assert len(name) <= 31
        assert name.endswith("_Vi")

    def test_make_translated_sheet_name_collision(self):
        existing = {"Sheet1_Vi"}
        name = _make_translated_sheet_name("Sheet1", "Vietnamese", existing)
        assert name != "Sheet1_Vi"
        assert name.endswith("_Vi")
        assert len(name) <= 31

    def test_31_char_limit_with_long_original(self):
        long_name = "VeryLongSheetNameThatExceeds"  # 27 chars + "_Vi" = 30 OK
        name = _make_translated_sheet_name(long_name, "Vietnamese", set())
        assert len(name) <= 31
        assert name.endswith("_Vi")

    def test_long_name_workbook(self, tmp_path: Path):
        src = tmp_path / "long.xlsx"
        _create_long_name_workbook(src)

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="en",
            target_lang="Vietnamese",
        )
        out = proc.process(src)
        wb = openpyxl.load_workbook(str(out))

        # All sheet names must be <= 31 chars
        for name in wb.sheetnames:
            assert len(name) <= 31, f"Sheet name '{name}' exceeds 31 chars"

        # Should have the translated version
        translated_sheets = [n for n in wb.sheetnames if n.endswith("_Vi")]
        assert len(translated_sheets) == 1
