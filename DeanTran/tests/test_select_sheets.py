"""
tests/test_select_sheets.py – Tests for SelectSheetsDialog logic.

These tests validate the dialog's helper functions and the
ExcelProcessor integration with sheet selection, without requiring
a running Qt event loop for most tests.
"""
from __future__ import annotations

import sys
from pathlib import Path
from unittest.mock import patch

import pytest
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.core.excel_processor import ExcelProcessor
from app.core.translators.translator_service import MockTranslator
from app.ui.dialogs.select_sheets_dialog import (
    collect_all_sheets,
    _load_saved_selections,
    _save_selections,
)


# ═══════════════════════════════════════════════════════════════════
# 1. collect_all_sheets
# ═══════════════════════════════════════════════════════════════════

class TestCollectAllSheets:
    def test_reads_sheets_from_file(self, tmp_path: Path):
        src = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Alpha"
        wb.create_sheet("Beta")
        wb.create_sheet("Gamma")
        wb.save(str(src))

        result = collect_all_sheets([str(src)])
        assert str(src) in result
        assert result[str(src)] == ["Alpha", "Beta", "Gamma"]

    def test_invalid_file_returns_empty_list(self, tmp_path: Path):
        fake = str(tmp_path / "nonexistent.xlsx")
        result = collect_all_sheets([fake])
        assert result[fake] == []

    def test_multiple_files(self, tmp_path: Path):
        f1 = tmp_path / "a.xlsx"
        wb1 = openpyxl.Workbook()
        wb1.active.title = "S1"
        wb1.save(str(f1))

        f2 = tmp_path / "b.xlsx"
        wb2 = openpyxl.Workbook()
        wb2.active.title = "X1"
        wb2.create_sheet("X2")
        wb2.save(str(f2))

        result = collect_all_sheets([str(f1), str(f2)])
        assert result[str(f1)] == ["S1"]
        assert result[str(f2)] == ["X1", "X2"]


# ═══════════════════════════════════════════════════════════════════
# 2. Persistence
# ═══════════════════════════════════════════════════════════════════

class TestPersistence:
    def test_save_and_load(self):
        test_data = {"/path/to/file.xlsx": ["Sheet1", "Sheet3"]}
        _save_selections(test_data)
        loaded = _load_saved_selections()
        assert loaded["/path/to/file.xlsx"] == ["Sheet1", "Sheet3"]

    def test_load_empty_returns_dict(self):
        # Clear any saved data
        _save_selections({})
        loaded = _load_saved_selections()
        assert isinstance(loaded, dict)


# ═══════════════════════════════════════════════════════════════════
# 3. Only selected sheets translated
# ═══════════════════════════════════════════════════════════════════

class TestSelectedSheetsTranslation:
    def test_only_selected_sheet_gets_translated(self, tmp_path: Path):
        src = tmp_path / "select.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Hello"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "World"
        ws3 = wb.create_sheet("Sheet3")
        ws3["A1"] = "Foo"
        wb.save(str(src))

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="en",
            target_lang="Vietnamese",
            sheets_to_translate=["Sheet2"],
        )
        out = proc.process(src)
        wb_out = openpyxl.load_workbook(str(out))

        assert "Sheet2_Vi" in wb_out.sheetnames
        assert "Sheet1_Vi" not in wb_out.sheetnames
        assert "Sheet3_Vi" not in wb_out.sheetnames

        # Original sheets all present
        assert "Sheet1" in wb_out.sheetnames
        assert "Sheet2" in wb_out.sheetnames
        assert "Sheet3" in wb_out.sheetnames

    def test_select_all_translates_all(self, tmp_path: Path):
        src = tmp_path / "all.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "A"
        ws1["A1"] = "Text1"
        wb.create_sheet("B")["A1"] = "Text2"
        wb.save(str(src))

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="en",
            target_lang="Vietnamese",
            sheets_to_translate=["A", "B"],  # All selected
        )
        out = proc.process(src)
        wb_out = openpyxl.load_workbook(str(out))

        assert "A_Vi" in wb_out.sheetnames
        assert "B_Vi" in wb_out.sheetnames

    def test_none_means_all(self, tmp_path: Path):
        """sheets_to_translate=None should translate every sheet."""
        src = tmp_path / "none.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "X"
        wb["X"]["A1"] = "Data"
        wb.create_sheet("Y")["A1"] = "Data2"
        wb.save(str(src))

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="en",
            target_lang="Vietnamese",
            sheets_to_translate=None,
        )
        out = proc.process(src)
        wb_out = openpyxl.load_workbook(str(out))
        assert "X_Vi" in wb_out.sheetnames
        assert "Y_Vi" in wb_out.sheetnames
