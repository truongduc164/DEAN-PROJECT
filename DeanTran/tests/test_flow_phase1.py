"""
tests/test_flow_phase1.py – Automated tests for Phase-1 core logic.

Covers:
  1. EventManager subscribe/emit
  2. MockTranslator deterministic output
  3. ExcelProcessor end-to-end on a temp workbook
  4. Skip-empty and skip-numeric behaviour
  5. build_output_path naming + no overwrite
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

# Ensure project root is on sys.path so "app.core…" imports work.
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from app.core.event_manager import EventManager, attach_console
from app.core.translators.translator_service import MockTranslator
from app.core.excel_processor import ExcelProcessor, build_output_path


# ── helpers ──────────────────────────────────────────────────────────

def _create_sample_workbook(path: Path) -> None:
    """Create a tiny .xlsx with mixed content for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    ws["A2"] = "World"
    ws["A3"] = 12345          # numeric – should be skipped
    ws["A4"] = ""             # empty – should be skipped
    ws["A5"] = "Translate me"
    ws["B1"] = "42.5"         # numeric string – should be skipped
    ws["B2"] = "Alpha"
    wb.save(str(path))


# ── EventManager tests ──────────────────────────────────────────────

class TestEventManager:
    def test_subscribe_and_emit(self):
        em = EventManager()
        received = []
        em.subscribe("log", lambda lvl, msg: received.append((lvl, msg)))
        em.log("INFO", "hello")
        assert received == [("INFO", "hello")]

    def test_multiple_subscribers(self):
        em = EventManager()
        a, b = [], []
        em.subscribe("log", lambda lvl, msg: a.append(msg))
        em.subscribe("log", lambda lvl, msg: b.append(msg))
        em.log("INFO", "x")
        assert a == ["x"] and b == ["x"]

    def test_progress_event(self):
        em = EventManager()
        values = []
        em.subscribe("progress", lambda cur, tot: values.append((cur, tot)))
        em.progress(3, 10)
        assert values == [(3, 10)]

    def test_error_event(self):
        em = EventManager()
        errors = []
        em.subscribe("error", lambda msg, exc: errors.append(msg))
        em.error("boom", ValueError("test"))
        assert errors == ["boom"]

    def test_subscriber_exception_does_not_crash(self):
        em = EventManager()
        em.subscribe("log", lambda *_: 1 / 0)  # will raise
        em.log("INFO", "safe")  # should NOT raise


# ── MockTranslator tests ────────────────────────────────────────────

class TestMockTranslator:
    def test_deterministic_output(self):
        t = MockTranslator()
        assert t.translate("hello", "en", "vi") == "[vi] hello"

    def test_call_count(self):
        t = MockTranslator()
        t.translate("a", "en", "vi")
        t.translate("b", "en", "vi")
        assert t.call_count == 2


# ── build_output_path tests ─────────────────────────────────────────

class TestBuildOutputPath:
    def test_vietnamese_suffix(self):
        p = build_output_path(Path("/data/Book1.xlsx"), "Vietnamese")
        assert p == Path("/data/Book1_Vi.xlsx")

    def test_english_suffix(self):
        p = build_output_path(Path("/data/Report.xlsx"), "English")
        assert p == Path("/data/Report_En.xlsx")

    def test_same_directory_as_input(self):
        p = build_output_path(Path("/some/folder/Test.xlsx"), "Vietnamese")
        assert p.parent == Path("/some/folder")

    def test_unknown_lang_uses_first_two_chars(self):
        p = build_output_path(Path("/data/x.xlsx"), "French")
        assert p == Path("/data/x_Fr.xlsx")


# ── ExcelProcessor tests ────────────────────────────────────────────

class TestExcelProcessor:
    def test_end_to_end(self, tmp_path: Path):
        """Translate a sample workbook and verify output."""
        src = tmp_path / "sample.xlsx"
        _create_sample_workbook(src)

        translator = MockTranslator()
        em = EventManager()
        logs: list[tuple[str, str]] = []
        em.subscribe("log", lambda lvl, msg: logs.append((lvl, msg)))

        proc = ExcelProcessor(
            translator=translator,
            event_manager=em,
            skip_empty=True,
            skip_numeric=True,
            source_lang="zh",
            target_lang="Vietnamese",
        )
        out_path = proc.process(src)

        # Output file must exist in SAME directory with _Vi suffix
        assert out_path.exists()
        assert out_path.parent == src.parent
        assert out_path.name == "sample_Vi.xlsx"

        # Original file must still exist and be unchanged
        assert src.exists()
        orig_wb = openpyxl.load_workbook(str(src))
        assert orig_wb.active["A1"].value == "Hello"

        # Read output and verify TRANSLATED sheets
        wb = openpyxl.load_workbook(str(out_path))
        assert "Sheet1" in wb.sheetnames      # original kept
        assert "Sheet1_Vi" in wb.sheetnames    # translated copy created

        # Original sheet should be unchanged
        ws_orig = wb["Sheet1"]
        assert ws_orig["A1"].value == "Hello"
        assert ws_orig["A2"].value == "World"

        # Translated sheet should have mock translations
        ws_trans = wb["Sheet1_Vi"]
        assert ws_trans["A1"].value == "[Vietnamese] Hello"
        assert ws_trans["A2"].value == "[Vietnamese] World"
        # Numeric cell kept as-is
        assert ws_trans["A3"].value == 12345
        # Empty cell stays empty
        assert ws_trans["A4"].value is None or ws_trans["A4"].value == ""
        # Text cell translated
        assert ws_trans["A5"].value == "[Vietnamese] Translate me"
        # Numeric string skipped
        assert ws_trans["B1"].value == "42.5"
        # Text translated
        assert ws_trans["B2"].value == "[Vietnamese] Alpha"

        # Translator should have been called exactly for eligible cells
        # Eligible: A1, A2, A5, B2 = 4
        assert translator.call_count == 4

        # Logs should exist
        assert any("Loading workbook" in msg for _, msg in logs)

    def test_skip_numeric_disabled(self, tmp_path: Path):
        """When skip_numeric=False, numeric strings get translated too."""
        src = tmp_path / "sample.xlsx"
        _create_sample_workbook(src)

        translator = MockTranslator()
        proc = ExcelProcessor(
            translator=translator,
            skip_empty=True,
            skip_numeric=False,
            source_lang="zh",
            target_lang="Vietnamese",
        )
        out_path = proc.process(src)
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Sheet1_Vi"]
        # "42.5" is a string so it should be translated when skip_numeric=False
        assert ws["B1"].value == "[Vietnamese] 42.5"
        # Pure int 12345 should also be translated (converted to str then translated)
        assert ws["A3"].value == "[Vietnamese] 12345"

    def test_no_overwrite_original(self, tmp_path: Path):
        """Original file must NEVER be overwritten."""
        src = tmp_path / "sample.xlsx"
        _create_sample_workbook(src)

        # Record original content
        orig_data = src.read_bytes()

        proc = ExcelProcessor(
            translator=MockTranslator(),
            source_lang="zh",
            target_lang="Vietnamese",
        )
        out_path = proc.process(src)

        # Output path must differ from input path
        assert out_path != src
        # Original file bytes must be unchanged
        assert src.read_bytes() == orig_data
