"""
tests/test_phase2.py – Phase-2 tests for Prompt, Glossary, Regex, Pipeline.

Covers:
  1. PromptStore load/save per mode
  2. GlossaryLoader + TermOverride correctness
  3. RegexProtection round-trip (protect → mock translate → restore)
  4. Full ExcelProcessor pipeline with glossary + protection + prompt
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from app.storage.prompt_store import PromptStore
from app.term_engine.glossary_loader import GlossaryLoader
from app.term_engine.term_override import TermOverride
from app.term_engine import regex_protection
from app.core.event_manager import EventManager
from app.core.translators.translator_service import MockTranslator
from app.core.excel_processor import ExcelProcessor


# ═══════════════════════════════════════════════════════════════════
# 1. PromptStore
# ═══════════════════════════════════════════════════════════════════

class TestPromptStore:
    def test_defaults_exist(self, tmp_path: Path):
        store = PromptStore(config_dir=tmp_path)
        # Built-in defaults should be available
        assert "SOP" in store.modes()
        assert "General" in store.modes()
        prompt = store.get("SOP")
        assert len(prompt) > 10  # non-trivial default

    def test_get_unknown_mode_returns_empty(self, tmp_path: Path):
        store = PromptStore(config_dir=tmp_path)
        assert store.get("NonExistent") == ""

    def test_set_and_get(self, tmp_path: Path):
        store = PromptStore(config_dir=tmp_path)
        store.set("CustomMode", "My custom prompt text")
        assert store.get("CustomMode") == "My custom prompt text"

    def test_save_and_reload(self, tmp_path: Path):
        store = PromptStore(config_dir=tmp_path)
        store.set("MyMode", "Prompt A")
        store.save()

        # Create a fresh instance that loads from the same file
        store2 = PromptStore(config_dir=tmp_path)
        assert store2.get("MyMode") == "Prompt A"

    def test_overwrite_default(self, tmp_path: Path):
        store = PromptStore(config_dir=tmp_path)
        original = store.get("SOP")
        store.set("SOP", "Overridden prompt")
        store.save()

        store2 = PromptStore(config_dir=tmp_path)
        assert store2.get("SOP") == "Overridden prompt"
        assert store2.get("SOP") != original

    def test_delete_mode(self, tmp_path: Path):
        store = PromptStore(config_dir=tmp_path)
        store.set("Temp", "temp prompt")
        store.delete("Temp")
        # Falls back to default (empty for unknown mode)
        assert store.get("Temp") == ""

    def test_modes_list(self, tmp_path: Path):
        store = PromptStore(config_dir=tmp_path)
        store.set("Alpha", "a")
        store.set("Beta", "b")
        modes = store.modes()
        assert "Alpha" in modes
        assert "Beta" in modes
        assert "SOP" in modes  # default


# ═══════════════════════════════════════════════════════════════════
# 2. Glossary & TermOverride
# ═══════════════════════════════════════════════════════════════════

def _write_glossary(path: Path, entries: dict) -> None:
    path.write_text(json.dumps(entries, ensure_ascii=False), encoding="utf-8")


class TestGlossaryLoader:
    def test_load_json(self, tmp_path: Path):
        f = tmp_path / "glossary.json"
        _write_glossary(f, {"BOM": "Bảng vật tư", "SMT": "Công nghệ dán bề mặt"})
        gl = GlossaryLoader(f)
        assert gl.lookup("BOM") == "Bảng vật tư"
        assert gl.lookup("bom") == "Bảng vật tư"  # case-insensitive
        assert gl.lookup("SMT") == "Công nghệ dán bề mặt"
        assert gl.lookup("unknown") is None

    def test_empty_when_no_file(self, tmp_path: Path):
        gl = GlossaryLoader(tmp_path / "missing.json")
        assert gl.is_empty()

    def test_empty_when_none(self):
        gl = GlossaryLoader(None)
        assert gl.is_empty()


class TestTermOverride:
    def _make_override(self, tmp_path: Path, entries: dict) -> TermOverride:
        f = tmp_path / "glossary.json"
        _write_glossary(f, entries)
        gl = GlossaryLoader(f)
        return TermOverride(gl)

    def test_pre_replace(self, tmp_path: Path):
        to = self._make_override(tmp_path, {"BOM": "Bảng vật tư"})
        result = to.pre_replace("Check BOM before production")
        assert "Bảng vật tư" in result
        assert "BOM" not in result

    def test_post_fix(self, tmp_path: Path):
        to = self._make_override(tmp_path, {"SMT": "Công nghệ dán"})
        # Suppose the LLM output still contains the source term
        result = to.post_fix("Apply SMT technology")
        assert "Công nghệ dán" in result

    def test_case_insensitive(self, tmp_path: Path):
        to = self._make_override(tmp_path, {"QC": "Kiểm tra chất lượng"})
        result = to.pre_replace("Run qc check")
        assert "Kiểm tra chất lượng" in result

    def test_longer_term_wins(self, tmp_path: Path):
        to = self._make_override(tmp_path, {
            "SMT": "Dán bề mặt",
            "SMT line": "Dây chuyền SMT",
        })
        result = to.pre_replace("The SMT line is busy")
        assert "Dây chuyền SMT" in result

    def test_no_glossary_passthrough(self):
        gl = GlossaryLoader(None)
        to = TermOverride(gl)
        assert to.pre_replace("unchanged") == "unchanged"


# ═══════════════════════════════════════════════════════════════════
# 3. Regex Protection
# ═══════════════════════════════════════════════════════════════════

class TestRegexProtection:
    def test_product_codes(self):
        text = "Use BTN-8 and SMI-13 for assembly"
        protected, mapping = regex_protection.protect(text)
        # Codes should be replaced with placeholders
        assert "BTN-8" not in protected
        assert "SMI-13" not in protected
        # Placeholders present (PUA chars)
        assert "\uE000\uE001" in protected
        # Round-trip restore
        restored = regex_protection.restore(protected, mapping)
        assert "BTN-8" in restored
        assert "SMI-13" in restored

    def test_ratios(self):
        text = "Mix ratio is 5:1:1.2"
        protected, mapping = regex_protection.protect(text)
        assert "5:1:1.2" not in protected
        restored = regex_protection.restore(protected, mapping)
        assert "5:1:1.2" in restored

    def test_pantone_code(self):
        text = "Color is Pantone 471C"
        protected, mapping = regex_protection.protect(text)
        assert "471C" not in protected
        restored = regex_protection.restore(protected, mapping)
        assert "471C" in restored

    def test_units(self):
        text = "Size is 12mm x 3.5cm at 25°C"
        protected, mapping = regex_protection.protect(text)
        assert "12mm" not in protected
        assert "25°C" not in protected
        restored = regex_protection.restore(protected, mapping)
        assert "12mm" in restored
        assert "25°C" in restored

    def test_pure_numbers(self):
        text = "Value is 456"
        protected, mapping = regex_protection.protect(text)
        assert "456" not in protected
        restored = regex_protection.restore(protected, mapping)
        assert "456" in restored

    def test_serial_ids(self):
        text = "Part XDK2512016 and YB0807944"
        protected, mapping = regex_protection.protect(text)
        assert "XDK2512016" not in protected
        assert "YB0807944" not in protected
        restored = regex_protection.restore(protected, mapping)
        assert "XDK2512016" in restored
        assert "YB0807944" in restored

    def test_round_trip_with_mock_translator(self):
        """Full round-trip: protect -> mock translate -> restore."""
        text = "Apply BTN-8 at 25°C with ratio 5:1:1.2"
        protected, mapping = regex_protection.protect(text)

        # Simulate mock translation (wraps with [vi])
        translated = f"[vi] {protected}"

        restored = regex_protection.restore(translated, mapping)
        assert "BTN-8" in restored
        assert "25°C" in restored
        assert "5:1:1.2" in restored

    def test_no_tokens_passthrough(self):
        text = "Just plain text"
        protected, mapping = regex_protection.protect(text)
        assert protected == text
        assert mapping == {}

    def test_dates(self):
        text = "Date: 2024-01-15"
        protected, mapping = regex_protection.protect(text)
        assert "2024-01-15" not in protected
        restored = regex_protection.restore(protected, mapping)
        assert "2024-01-15" in restored


# ═══════════════════════════════════════════════════════════════════
# 4. Integrated Excel pipeline (prompt + glossary + regex)
# ═══════════════════════════════════════════════════════════════════

def _create_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Check BOM status"
    ws["A2"] = "Part BTN-8 at 25°C"
    ws["A3"] = 42                     # numeric – skip
    ws["A4"] = "Apply SMT process"
    ws["B1"] = ""                     # empty – skip
    wb.save(str(path))


class TestIntegratedPipeline:
    def test_pipeline_with_glossary_and_protection(self, tmp_path: Path):
        # Setup glossary
        gf = tmp_path / "glossary.json"
        _write_glossary(gf, {"BOM": "Bảng vật tư", "SMT": "Công nghệ dán"})
        glossary = GlossaryLoader(gf)

        # Setup workbook
        src = tmp_path / "input.xlsx"
        _create_xlsx(src)

        em = EventManager()
        logs: list = []
        em.subscribe("log", lambda lvl, msg: logs.append(msg))

        proc = ExcelProcessor(
            translator=MockTranslator(),
            event_manager=em,
            skip_empty=True,
            skip_numeric=True,
            source_lang="zh",
            target_lang="Vietnamese",
            glossary=glossary,
            use_glossary=True,
            use_protection=True,
            prompt="Translate as SOP",
        )
        out_path = proc.process(src)
        assert out_path.exists()
        assert out_path.name == "input_Vi.xlsx"

        wb = openpyxl.load_workbook(str(out_path))

        # Both original and translated sheets should exist
        assert "Data" in wb.sheetnames
        assert "Data_Vi" in wb.sheetnames

        ws = wb["Data_Vi"]  # Check translated sheet

        # A1 – "Check BOM status": BOM pre-replaced → "Check Bảng vật tư status"
        # then translated [Vietnamese], then post-fixed
        a1 = ws["A1"].value
        assert "Bảng vật tư" in a1

        # A2 – "Part BTN-8 at 25°C": codes protected, translated, restored
        a2 = ws["A2"].value
        assert "BTN-8" in a2       # restored by regex protection
        assert "25°C" in a2        # restored by regex protection

        # A3 – numeric, should be unchanged
        assert ws["A3"].value == 42

        # A4 – "Apply SMT process": SMT replaced by glossary
        a4 = ws["A4"].value
        assert "Công nghệ dán" in a4

        # Logs should mention glossary and protection
        all_logs = " ".join(logs)
        assert "Glossary override: ENABLED" in all_logs
        assert "Regex protection: ENABLED" in all_logs

    def test_pipeline_without_glossary(self, tmp_path: Path):
        src = tmp_path / "input.xlsx"
        _create_xlsx(src)

        proc = ExcelProcessor(
            translator=MockTranslator(),
            skip_empty=True,
            skip_numeric=True,
            use_glossary=False,
            use_protection=False,
        )
        out_path = proc.process(src)
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Data_Vi"]  # Check translated sheet
        # Without glossary, BOM stays as literal
        assert "BOM" in ws["A1"].value
